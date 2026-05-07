from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from drivers.dependencies import require_module_access
from drivers.db_client import get_db
from sqlalchemy.orm import Session
from sqlalchemy import text
import json
import urllib.request
import urllib.error
from typing import Optional
from openpyxl import load_workbook
from datetime import datetime, date
import os
import time


router = APIRouter()


SALEWORK_CLIENT_ID = os.getenv("SALEWORK_CLIENT_ID", "2573")
SALEWORK_TOKEN = os.getenv("SALEWORK_TOKEN", "/Kx80W61t30ZsEKFXJB4svDkmz98zLZ6Wpkg1V82UNWDRBSzKv7B04cCB1RSPYf5")
SALEWORK_PRODUCT_LIST_URL = "https://salework.net/api/open/stock/v1/product/list"
SALEWORK_PRODUCT_REPORT_URL = "https://salework.net/api/open/stock/v1/report/product"

def _current_period_month() -> str:
    return datetime.now().strftime("%Y-%m")


def _get_active_period_month(db: Session) -> str:
    # Prefer periods table if exists; fallback to current month
    try:
        row = db.execute(
            text("""
                SELECT period_month
                FROM inventory_check_periods
                WHERE is_active = 1
                ORDER BY id DESC
                LIMIT 1
            """)
        ).fetchone()
        if row and row[0]:
            return str(row[0])
    except Exception:
        pass
    return _current_period_month()


@router.get("/inventory-check/periods")
def list_inventory_check_periods(
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check")),
):
    """
    Danh sách kỳ kiểm tồn (period_month) + kỳ đang active.
    """
    try:
        rows = db.execute(
            text("""
                SELECT period_month, is_active, closed_at
                FROM inventory_check_periods
                ORDER BY period_month DESC
                LIMIT 24
            """)
        ).fetchall()
        items = [
            {
                "period_month": r[0],
                "is_active": bool(r[1]),
                "closed_at": r[2].strftime("%Y-%m-%d %H:%M:%S") if r[2] else None,
            }
            for r in rows
        ]
        active = next((x["period_month"] for x in items if x["is_active"]), _current_period_month())
        return {"status": "success", "active_period_month": active, "items": items}
    except Exception:
        # If table doesn't exist yet
        return {"status": "success", "active_period_month": _current_period_month(), "items": []}


def _fetch_salework_product_list() -> dict:
    req = urllib.request.Request(
        SALEWORK_PRODUCT_LIST_URL,
        method="GET",
        headers={
            "client-id": SALEWORK_CLIENT_ID,
            "token": SALEWORK_TOKEN,
        },
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw)
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore") if hasattr(e, "read") else str(e)
        raise HTTPException(status_code=502, detail=f"Salework HTTPError: {detail}")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Salework fetch failed: {str(e)}")


def _fetch_salework_product_report(time_start: int, time_end: int) -> dict:
    req = urllib.request.Request(
        SALEWORK_PRODUCT_REPORT_URL,
        method="POST",
        headers={
            "client-id": SALEWORK_CLIENT_ID,
            "token": SALEWORK_TOKEN,
            "Content-Type": "application/json",
        },
        data=json.dumps({"time_start": time_start, "time_end": time_end}).encode("utf-8"),
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as resp:
            raw = resp.read().decode("utf-8")
            return json.loads(raw)
    except urllib.error.HTTPError as e:
        detail = e.read().decode("utf-8", errors="ignore") if hasattr(e, "read") else str(e)
        raise HTTPException(status_code=502, detail=f"Salework HTTPError: {detail}")
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Salework fetch failed: {str(e)}")


@router.post("/inventory-check/accounting/init-openings-from-salework")
def init_accounting_openings_from_salework(
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check", require_manage=True)),
):
    """
    Khởi động kiểm tồn (ONE-TIME cho toàn hệ thống):
    - Lấy tồn hiện tại từ Salework
    - Gán vào "Tồn kế toán" (opening) cho kỳ active
    - Chỉ chạy 1 lần duy nhất: nếu đã khởi tạo rồi thì sẽ không cho chạy lại nữa.
    """
    period = _get_active_period_month(db)

    try:
        st = db.execute(
            text("SELECT initialized, initialized_at, initialized_period_month FROM inventory_check_bootstrap_state WHERE id = 1"),
        ).fetchone()
        if st and int(st[0] or 0) == 1:
            return {
                "status": "success",
                "period_month": period,
                "initialized": False,
                "already_initialized": True,
                "initialized_at": st[1].strftime("%Y-%m-%d %H:%M:%S") if st[1] else None,
                "initialized_period_month": st[2],
                "message": "Đã khởi tạo tồn kế toán (one-time) trước đó",
            }
    except Exception:
        # Nếu bảng chưa tồn tại / lỗi schema thì để lỗi rõ ràng cho user thay vì im lặng.
        raise

    payload = _fetch_salework_product_list()
    if payload.get("status") != "success":
        raise HTTPException(status_code=502, detail=f"Salework status != success: {payload.get('status')}")

    data = payload.get("data") or {}
    products = data.get("products") or {}
    if not isinstance(products, dict):
        raise HTTPException(status_code=502, detail="Unexpected Salework payload shape: data.products is not object")

    upsert_sql = text(
        """
        INSERT INTO accounting_stock_openings_v2 (period_month, product_code, opening_qty)
        VALUES (:p, :code, :qty)
        -- One-time bootstrap should overwrite stale/negative openings if any existed
        ON DUPLICATE KEY UPDATE opening_qty = VALUES(opening_qty)
        """
    )

    inserted = 0
    for code_key, p in products.items():
        p_code = (p.get("code") or code_key or "").strip()
        if not p_code:
            continue

        total = 0.0
        stocks = p.get("stocks") or []
        if isinstance(stocks, list):
            for s in stocks:
                try:
                    total += float((s or {}).get("value") or 0)
                except Exception:
                    pass

        # Giảm tải DB: chỉ khởi tạo các mã có tồn > 0
        if total <= 0:
            continue

        db.execute(upsert_sql, {"p": period, "code": p_code, "qty": total})
        inserted += 1

    db.execute(
        text(
            """
            INSERT INTO inventory_check_bootstrap_state (id, initialized, initialized_at, initialized_period_month, note)
            VALUES (1, 1, NOW(), :p, 'init-openings-from-salework')
            ON DUPLICATE KEY UPDATE
                initialized = VALUES(initialized),
                initialized_at = VALUES(initialized_at),
                initialized_period_month = VALUES(initialized_period_month),
                note = VALUES(note)
            """
        ),
        {"p": period},
    )
    db.commit()
    return {"status": "success", "period_month": period, "initialized": True, "already_initialized": False, "inserted": inserted}


@router.get("/inventory-check/accounting/bootstrap-state")
def get_inventory_check_bootstrap_state(
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check")),
):
    """
    Trạng thái khởi tạo one-time cho kiểm tồn.
    """
    try:
        st = db.execute(
            text("SELECT initialized, initialized_at, initialized_period_month, note FROM inventory_check_bootstrap_state WHERE id = 1")
        ).fetchone()
        if not st:
            return {"status": "success", "initialized": False}
        return {
            "status": "success",
            "initialized": bool(st[0]),
            "initialized_at": st[1].strftime("%Y-%m-%d %H:%M:%S") if st[1] else None,
            "initialized_period_month": st[2],
            "note": st[3],
        }
    except Exception:
        # If table doesn't exist yet
        return {"status": "success", "initialized": False}


@router.get("/inventory-check/salework/products")
def get_salework_products(
    page: int = 1,
    limit: int = 100,
    include_zero: bool = False,
    search: Optional[str] = None,
    user: dict = Depends(require_module_access("inventory-check")),
):
    """
    Proxy endpoint for realtime Salework stock.
    - include_zero=false: trả về chỉ sản phẩm có tồn > 0
    - search: lọc theo code/name (case-insensitive)
    - page/limit: phân trang server-side (để UI hiển thị mượt khi vài nghìn mã)
    """
    if page < 1:
        page = 1
    if limit < 1:
        limit = 1
    if limit > 500:
        limit = 500

    payload = _fetch_salework_product_list()
    if payload.get("status") != "success":
        raise HTTPException(status_code=502, detail=f"Salework status != success: {payload.get('status')}")

    data = payload.get("data") or {}
    products = data.get("products") or {}
    if not isinstance(products, dict):
        raise HTTPException(status_code=502, detail="Unexpected Salework payload shape: data.products is not object")

    q = (search or "").strip().lower()

    items = []
    for code, p in products.items():
        # thực tế code nằm trong p['code'], nhưng fallback theo key dict để an toàn
        p_code = (p.get("code") or code or "").strip()
        name = p.get("name") or ""
        stocks = p.get("stocks") or []

        total = 0.0
        if isinstance(stocks, list):
            for s in stocks:
                try:
                    total += float((s or {}).get("value") or 0)
                except Exception:
                    pass

        if not include_zero and total <= 0:
            continue

        if q:
            if q not in p_code.lower() and q not in str(name).lower():
                continue

        items.append(
            {
                "code": p_code,
                "name": name,
                "total_stock": total,
            }
        )

    items.sort(key=lambda x: x["code"] or "")
    total_count = len(items)
    start = (page - 1) * limit
    end = start + limit
    paged = items[start:end]
    return {
        "status": "success",
        "total": total_count,
        "page": page,
        "limit": limit,
        "count": len(paged),
        "items": paged,
    }


_INC_TYPES = {
    "Nhập mua hàng",
    "Nhập gia công",
    "Hàng hoàn",
    "Điều chỉnh tăng",
}
_DEC_TYPES = {
    "Xuất bán",
    "Xuất gia công",
    "Hủy hàng",
    "Điều chỉnh giảm",
    "Xuất nội bộ",
}


def _parse_excel_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    # string fallback
    try:
        s = str(v).strip()
        # accept YYYY-MM-DD or DD/MM/YYYY
        if "-" in s:
            return datetime.strptime(s[:10], "%Y-%m-%d").date()
        if "/" in s:
            return datetime.strptime(s[:10], "%d/%m/%Y").date()
    except Exception:
        return None
    return None


@router.post("/inventory-check/accounting/import")
def import_accounting_movements(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check", require_manage=True)),
):
    """
    Import biến động tăng/giảm cho cột Kế toán từ file Excel (bao_cao_ton_kho.xlsx).
    Map theo cột:
    - Ngày, Mã hàng, Loại biến động, Số lượng, Lý do, Bộ phận, Người phụ trách, Chứng từ, Ghi chú.
    """
    try:
        wb = load_workbook(file.file, data_only=True)
        ws = wb[wb.sheetnames[0]]

        # tìm header row chứa "Mã hàng" + "Loại biến động"
        header_row = None
        for r in range(1, 50):
            row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
            txt = " | ".join([str(x) for x in row_vals if x is not None])
            if "Mã hàng" in txt and "Loại biến động" in txt and "Số lượng" in txt:
                header_row = r
                break
        if not header_row:
            raise HTTPException(status_code=400, detail="Không tìm thấy header (cần có cột: Mã hàng, Loại biến động, Số lượng).")

        # map column index by header text
        col_map = {}
        for c in range(1, 30):
            v = ws.cell(row=header_row, column=c).value
            if not v:
                continue
            name = str(v).strip()
            col_map[name] = c

        def _col(*names):
            for n in names:
                if n in col_map:
                    return col_map[n]
            return None

        c_date = _col("Ngày")
        c_code = _col("Mã hàng")
        c_type = _col("Loại biến động")
        c_qty = _col("Số lượng")
        c_reason = _col("Lý do")
        c_dept = _col("Bộ phận")
        c_owner = _col("Người phụ trách")
        c_doc = _col("Chứng từ")
        c_note = _col("Ghi chú")

        if not (c_code and c_type and c_qty):
            raise HTTPException(status_code=400, detail="Thiếu cột bắt buộc: Mã hàng / Loại biến động / Số lượng.")

        inserted = 0
        skipped = 0
        unknown_types = {}

        insert_sql = text(
            """
            INSERT INTO accounting_stock_movements
                (period_month, product_code, movement_date, movement_type, direction, quantity, reason, department, owner, document_ref, note, source_file)
            VALUES
                (:period, :code, :d, :t, :dir, :q, :reason, :dept, :owner, :doc, :note, :src)
            """
        )
        period_month = _get_active_period_month(db)

        # data starts from next row
        for r in range(header_row + 1, ws.max_row + 1):
            code = ws.cell(row=r, column=c_code).value
            mtype = ws.cell(row=r, column=c_type).value
            qty = ws.cell(row=r, column=c_qty).value

            if code is None and mtype is None and qty is None:
                continue

            code_str = str(code).strip() if code is not None else ""
            if not code_str:
                skipped += 1
                continue

            mtype_str = str(mtype).strip() if mtype is not None else ""
            if not mtype_str:
                skipped += 1
                continue

            # direction mapping
            if mtype_str in _INC_TYPES:
                direction = "inc"
            elif mtype_str in _DEC_TYPES:
                direction = "dec"
            else:
                unknown_types[mtype_str] = unknown_types.get(mtype_str, 0) + 1
                skipped += 1
                continue

            try:
                q = float(qty or 0)
            except Exception:
                skipped += 1
                continue
            if q == 0:
                skipped += 1
                continue

            d = _parse_excel_date(ws.cell(row=r, column=c_date).value) if c_date else None
            reason = str(ws.cell(row=r, column=c_reason).value).strip() if c_reason and ws.cell(row=r, column=c_reason).value is not None else None
            dept = str(ws.cell(row=r, column=c_dept).value).strip() if c_dept and ws.cell(row=r, column=c_dept).value is not None else None
            owner = str(ws.cell(row=r, column=c_owner).value).strip() if c_owner and ws.cell(row=r, column=c_owner).value is not None else None
            doc = str(ws.cell(row=r, column=c_doc).value).strip() if c_doc and ws.cell(row=r, column=c_doc).value is not None else None
            note = str(ws.cell(row=r, column=c_note).value).strip() if c_note and ws.cell(row=r, column=c_note).value is not None else None

            db.execute(
                insert_sql,
                {
                    "period": period_month,
                    "code": code_str,
                    "d": d,
                    "t": mtype_str,
                    "dir": direction,
                    "q": q,
                    "reason": reason,
                    "dept": dept,
                    "owner": owner,
                    "doc": doc,
                    "note": note,
                    "src": file.filename,
                },
            )
            inserted += 1

        db.commit()

        return {
            "status": "success",
            "inserted": inserted,
            "skipped": skipped,
            "unknown_types": unknown_types,
        }
    except HTTPException:
        raise
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/inventory-check/accounting/movements")
def get_accounting_movements(
    product_code: str,
    page: int = 1,
    limit: int = 50,
    period_month: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check")),
):
    if page < 1:
        page = 1
    if limit < 1:
        limit = 1
    if limit > 200:
        limit = 200

    period = (period_month or "").strip() or _get_active_period_month(db)
    total = db.execute(
        text("SELECT COUNT(*) FROM accounting_stock_movements WHERE product_code = :c AND period_month = :p"),
        {"c": product_code, "p": period},
    ).scalar() or 0

    offset = (page - 1) * limit
    rows = db.execute(
        text(
            """
            SELECT id, movement_date, movement_type, direction, quantity, reason, department, owner, document_ref, note, created_at
            FROM accounting_stock_movements
            WHERE product_code = :c AND period_month = :p
            ORDER BY COALESCE(movement_date, DATE(created_at)) DESC, id DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        {"c": product_code, "p": period, "limit": limit, "offset": offset},
    ).fetchall()

    items = [
        {
            "id": r[0],
            "date": r[1].strftime("%Y-%m-%d") if r[1] else None,
            "type": r[2],
            "direction": r[3],
            "quantity": float(r[4] or 0),
            "reason": r[5],
            "department": r[6],
            "owner": r[7],
            "document_ref": r[8],
            "note": r[9],
            "created_at": r[10].strftime("%Y-%m-%d %H:%M:%S") if r[10] else None,
        }
        for r in rows
    ]

    return {"status": "success", "total": int(total), "page": page, "limit": limit, "items": items}


@router.get("/inventory-check/accounting/summary")
def get_accounting_summary_for_codes(
    codes: str,
    period_month: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check")),
):
    """
    Trả về tổng tăng/tổng giảm theo danh sách code (comma-separated).
    Dùng để render cột Tăng/Giảm kế toán theo trang hiện tại mà không gọi N lần.
    """
    code_list = [c.strip() for c in (codes or "").split(",") if c.strip()]
    if not code_list:
        return {"status": "success", "items": {}}

    # tránh query quá dài
    code_list = code_list[:500]
    placeholders = ",".join([f":c{i}" for i in range(len(code_list))])
    params = {f"c{i}": code_list[i] for i in range(len(code_list))}

    period = (period_month or "").strip() or _get_active_period_month(db)
    rows = db.execute(
        text(
            f"""
            SELECT product_code,
                   SUM(CASE WHEN direction='inc' THEN quantity ELSE 0 END) AS inc_qty,
                   SUM(CASE WHEN direction='dec' THEN quantity ELSE 0 END) AS dec_qty
            FROM accounting_stock_movements
            WHERE period_month = :p AND product_code IN ({placeholders})
            GROUP BY product_code
            """
        ),
        {**params, "p": period},
    ).fetchall()

    out = {c: {"inc": 0.0, "dec": 0.0} for c in code_list}
    for code, inc_qty, dec_qty in rows:
        out[str(code)] = {"inc": float(inc_qty or 0), "dec": float(dec_qty or 0)}

    return {"status": "success", "items": out}

@router.get("/inventory-check/accounting/openings")
def get_accounting_openings_for_codes(
    codes: str,
    period_month: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check")),
):
    """
    Trả về tồn đầu kế toán theo danh sách code (comma-separated).
    """
    code_list = [c.strip() for c in (codes or "").split(",") if c.strip()]
    if not code_list:
        return {"status": "success", "items": {}}

    code_list = code_list[:500]
    placeholders = ",".join([f":c{i}" for i in range(len(code_list))])
    params = {f"c{i}": code_list[i] for i in range(len(code_list))}

    period = (period_month or "").strip() or _get_active_period_month(db)
    rows = db.execute(
        text(
            f"""
            SELECT product_code, opening_qty
            FROM accounting_stock_openings_v2
            WHERE period_month = :p AND product_code IN ({placeholders})
            """
        ),
        {**params, "p": period},
    ).fetchall()

    out = {c: 0.0 for c in code_list}
    for code, qty in rows:
        out[str(code)] = float(qty or 0)

    return {"status": "success", "items": out}


@router.post("/inventory-check/accounting/close")
def close_accounting_period(
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check", require_manage=True)),
):
    """
    "Chốt số theo kỳ":
    - Tồn cuối kỳ N = tồn đầu + tăng - giảm (lọc theo period_month active)
    - Tạo kỳ N+1 (next month) và ghi tồn đầu N+1 = tồn cuối N
    - Movements của kỳ N được GIỮ LẠI (không xóa)
    - Kỳ N+1 bắt đầu với movements rỗng
    """
    cur_period = _get_active_period_month(db)
    # next month
    y, m = cur_period.split("-")
    ny = int(y)
    nm = int(m) + 1
    if nm == 13:
        nm = 1
        ny += 1
    next_period = f"{ny:04d}-{nm:02d}"

    # Ensure periods table + switch active
    try:
        db.execute(text("INSERT IGNORE INTO inventory_check_periods (period_month, is_active) VALUES (:p, 0)"), {"p": next_period})
        db.execute(text("UPDATE inventory_check_periods SET is_active = 0, closed_at = IF(period_month=:cur, NOW(), closed_at) WHERE period_month = :cur"), {"cur": cur_period})
        db.execute(text("UPDATE inventory_check_periods SET is_active = 1 WHERE period_month = :nxt"), {"nxt": next_period})
    except Exception:
        # If periods table not ready, still proceed with openings using month strings
        pass

    # Lấy danh sách code cần chốt: union openings + movements của kỳ hiện tại
    codes = set()
    for (c,) in db.execute(text("SELECT product_code FROM accounting_stock_openings_v2 WHERE period_month = :p"), {"p": cur_period}).fetchall():
        if c:
            codes.add(str(c))
    for (c,) in db.execute(text("SELECT DISTINCT product_code FROM accounting_stock_movements WHERE period_month = :p"), {"p": cur_period}).fetchall():
        if c:
            codes.add(str(c))

    if not codes:
        # reset state cho kỳ mới (sales sync)
        try:
            db.execute(text("INSERT INTO inventory_check_sales_sync_state_v2 (period_month, last_sync_ms) VALUES (:p, 0) ON DUPLICATE KEY UPDATE last_sync_ms=0"), {"p": next_period})
        except Exception:
            pass
        db.commit()
        return {"status": "success", "updated": 0, "current_period_month": cur_period, "next_period_month": next_period}

    # Tính closing theo từng code bằng SQL aggregate
    # closing = opening + inc - dec
    # opening mặc định 0 nếu chưa có
    placeholders = ",".join([f":c{i}" for i in range(len(codes))])
    code_list = list(codes)
    params = {f"c{i}": code_list[i] for i in range(len(code_list))}

    rows = db.execute(
        text(
            f"""
            SELECT
                c.code AS product_code,
                COALESCE(o.opening_qty, 0) AS opening_qty,
                COALESCE(m.inc_qty, 0) AS inc_qty,
                COALESCE(m.dec_qty, 0) AS dec_qty
            FROM (
                SELECT :c0 AS code
                {"".join([f" UNION ALL SELECT :c{i}" for i in range(1, len(code_list))])}
            ) c
            LEFT JOIN accounting_stock_openings_v2 o ON o.product_code = c.code AND o.period_month = :cur_period
            LEFT JOIN (
                SELECT product_code,
                       SUM(CASE WHEN direction='inc' THEN quantity ELSE 0 END) AS inc_qty,
                       SUM(CASE WHEN direction='dec' THEN quantity ELSE 0 END) AS dec_qty
                FROM accounting_stock_movements
                WHERE period_month = :cur_period AND product_code IN ({placeholders})
                GROUP BY product_code
            ) m ON m.product_code = c.code
            """
        ),
        {**params, "cur_period": cur_period},
    ).fetchall()

    upsert_sql = text(
        """
        INSERT INTO accounting_stock_openings_v2 (period_month, product_code, opening_qty)
        VALUES (:period, :code, :qty)
        ON DUPLICATE KEY UPDATE opening_qty = VALUES(opening_qty)
        """
    )

    updated = 0
    for code, opening_qty, inc_qty, dec_qty in rows:
        closing = float(opening_qty or 0) + float(inc_qty or 0) - float(dec_qty or 0)
        db.execute(upsert_sql, {"period": next_period, "code": str(code), "qty": closing})
        updated += 1

    # Reset sales realtime state for next period (movements are kept by period)
    try:
        db.execute(text("INSERT INTO inventory_check_sales_sync_state_v2 (period_month, last_sync_ms) VALUES (:p, 0) ON DUPLICATE KEY UPDATE last_sync_ms=0"), {"p": next_period})
    except Exception:
        pass
    db.commit()

    return {"status": "success", "updated": updated, "current_period_month": cur_period, "next_period_month": next_period}


@router.post("/inventory-check/accounting/sync-sales")
def sync_accounting_sales_from_salework(
    time_start: int,
    time_end: int,
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check", require_manage=True)),
):
    """
    Sync giảm Kế toán từ số bán Salework (realtime).
    - Lấy report theo time_start/time_end (ms)
    - Ghi vào accounting_stock_movements với movement_type='Xuất bán', direction='dec'
    - Chống cộng trùng: replace toàn bộ dữ liệu theo document_ref của window.
    """
    if time_start <= 0 or time_end <= 0 or time_end <= time_start:
        raise HTTPException(status_code=400, detail="time_start/time_end không hợp lệ")

    payload = _fetch_salework_product_report(time_start, time_end)
    if payload.get("status") != "success":
        raise HTTPException(status_code=502, detail=f"Salework status != success: {payload.get('status')}")

    data = payload.get("data") or {}
    report = data.get("product_report") or []

    # Thực tế Salework có thể trả dict theo channel hoặc list; hỗ trợ cả 2.
    agg = {}  # code -> {qty, channels: {channel: qty}}

    def add(code: str, qty: float, channel: str):
        if not code:
            return
        row = agg.get(code)
        if not row:
            row = {"qty": 0.0, "channels": {}}
            agg[code] = row
        row["qty"] += float(qty or 0)
        ch = row["channels"].get(channel, 0.0)
        row["channels"][channel] = ch + float(qty or 0)

    if isinstance(report, dict):
        for channel, blocks in report.items():
            if not isinstance(blocks, list):
                continue
            for b in blocks:
                prods = (b or {}).get("products") or []
                if not isinstance(prods, list):
                    continue
                for p in prods:
                    code = (p or {}).get("code")
                    qty = (p or {}).get("amount") or 0
                    add(str(code).strip() if code else "", qty, str(channel))
    elif isinstance(report, list):
        # docs shape: list product_report
        for p in report:
            code = (p or {}).get("code")
            qty = (p or {}).get("amount") or 0
            add(str(code).strip() if code else "", qty, "Salework")
    else:
        raise HTTPException(status_code=502, detail="Unexpected product_report shape")

    # Replace theo window để không cộng trùng
    doc_ref = f"SW_SALES:{time_start}-{time_end}"
    db.execute(
        text(
            """
            DELETE FROM accounting_stock_movements
            WHERE movement_type = 'Xuất bán'
              AND document_ref = :doc
            """
        ),
        {"doc": doc_ref},
    )

    insert_sql = text(
        """
        INSERT INTO accounting_stock_movements
            (product_code, movement_date, movement_type, direction, quantity, reason, document_ref, note, source_file)
        VALUES
            (:code, :d, 'Xuất bán', 'dec', :q, :reason, :doc, :note, :src)
        """
    )

    d = datetime.now().date()
    inserted = 0
    for code, row in agg.items():
        q = float(row.get("qty") or 0)
        if q <= 0:
            continue
        note = json.dumps({"channels": row.get("channels") or {}}, ensure_ascii=False)
        db.execute(
            insert_sql,
            {
                "code": code,
                "d": d,
                "q": q,
                "reason": "Số bán Salework (realtime)",
                "doc": doc_ref,
                "note": note,
                "src": "salework_api",
            },
        )
        inserted += 1

    db.commit()
    return {"status": "success", "inserted_codes": inserted, "document_ref": doc_ref}


@router.post("/inventory-check/accounting/sync-sales-realtime")
def sync_accounting_sales_realtime(
    db: Session = Depends(get_db),
    user: dict = Depends(require_module_access("inventory-check", require_manage=True)),
):
    """
    Realtime sync "giảm" từ Salework theo cơ chế delta:
    - Tự đọc last_sync_ms trong DB
    - Gọi report/product từ last_sync_ms -> now
    - Insert các movement_type='Xuất bán' theo delta window (không replace tổng)
    - Update last_sync_ms = now
    """
    period = _get_active_period_month(db)
    # Ensure state row exists (per period)
    row = db.execute(
        text("SELECT last_sync_ms FROM inventory_check_sales_sync_state_v2 WHERE period_month = :p"),
        {"p": period},
    ).fetchone()
    if not row:
        db.execute(text("INSERT INTO inventory_check_sales_sync_state_v2 (period_month, last_sync_ms) VALUES (:p, 0)"), {"p": period})
        db.commit()
        last_sync_ms = 0
    else:
        last_sync_ms = int(row[0] or 0)

    now_ms = int(time.time() * 1000)
    if last_sync_ms <= 0:
        # bootstrap: chỉ lấy 5 phút gần nhất để tránh insert quá lớn lần đầu
        last_sync_ms = max(0, now_ms - 5 * 60 * 1000)

    if now_ms - last_sync_ms < 10_000:
        return {"status": "success", "inserted_codes": 0, "window": [last_sync_ms, now_ms]}

    payload = _fetch_salework_product_report(last_sync_ms, now_ms)
    if payload.get("status") != "success":
        raise HTTPException(status_code=502, detail=f"Salework status != success: {payload.get('status')}")

    data = payload.get("data") or {}
    report = data.get("product_report") or []

    agg = {}  # code -> qty

    def add(code: str, qty: float):
        if not code:
            return
        agg[code] = float(agg.get(code, 0)) + float(qty or 0)

    if isinstance(report, dict):
        for _, blocks in report.items():
            if not isinstance(blocks, list):
                continue
            for b in blocks:
                prods = (b or {}).get("products") or []
                if not isinstance(prods, list):
                    continue
                for p in prods:
                    code = (p or {}).get("code")
                    qty = (p or {}).get("amount") or 0
                    add(str(code).strip() if code else "", qty)
    elif isinstance(report, list):
        for p in report:
            code = (p or {}).get("code")
            qty = (p or {}).get("amount") or 0
            add(str(code).strip() if code else "", qty)
    else:
        raise HTTPException(status_code=502, detail="Unexpected product_report shape")

    # Upsert theo ngày để hạn chế phình dữ liệu:
    # 1 dòng / product_code / ngày / period_month cho movement_type='Xuất bán'
    day_str = datetime.now().strftime("%Y-%m-%d")
    doc_ref = f"SW_SALES_DAY:{day_str}"
    insert_sql = text(
        """
        INSERT INTO accounting_stock_movements
            (period_month, product_code, movement_date, movement_type, direction, quantity, reason, document_ref, source_file)
        VALUES
            (:period, :code, :day, 'Xuất bán', 'dec', :q, :reason, :doc, 'salework_api')
        ON DUPLICATE KEY UPDATE
            quantity = quantity + VALUES(quantity)
        """
    )

    inserted = 0
    for code, qty in agg.items():
        q = float(qty or 0)
        if q <= 0:
            continue
        db.execute(
            insert_sql,
            {
                "period": period,
                "code": code,
                "q": q,
                "day": day_str,
                "reason": "Số bán Salework (realtime delta)",
                "doc": doc_ref,
            },
        )
        inserted += 1

    db.execute(
        text("UPDATE inventory_check_sales_sync_state_v2 SET last_sync_ms = :ms WHERE period_month = :p"),
        {"ms": now_ms, "p": period},
    )
    db.commit()

    return {"status": "success", "inserted_codes": inserted, "window": [last_sync_ms, now_ms], "document_ref": doc_ref}


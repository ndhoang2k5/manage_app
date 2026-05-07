from email import header
from sqlalchemy.orm import Session
from sqlalchemy import text
from sqlalchemy.exc import IntegrityError
import json
from entities.production import BOMCreateRequest, ProductionOrderCreateRequest, QuickProductionRequest, ReceiveGoodsRequest, ProductionUpdateRequest, UpdateProgressRequest, ProgressItem, ProductionMaterialUpdateItem, ProductionSizeUpdateItem
from typing import List, Optional, Dict
from decimal import Decimal
import time
from io import BytesIO
from datetime import date, datetime

from openpyxl import Workbook
from copy import copy

class ProductionService:
    def __init__(self, db: Session):
        self.db = db

    # 1. Tạo Công thức (BOM) - GIỮ NGUYÊN
    def create_bom(self, data: BOMCreateRequest):
        try:
            query_bom = text("INSERT INTO bom (product_variant_id, name) VALUES (:pid, :name)")
            self.db.execute(query_bom, {"pid": data.product_variant_id, "name": data.name})
            bom_id = self.db.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]

            query_detail = text("""
                INSERT INTO bom_materials (bom_id, material_variant_id, quantity_needed)
                VALUES (:bid, :mid, :qty)
            """)
            for item in data.materials:
                self.db.execute(query_detail, {
                    "bid": bom_id, 
                    "mid": item.material_variant_id, 
                    "qty": item.quantity_needed
                })
            
            self.db.commit()
            return {"status": "success", "message": "Đã tạo công thức thành công"}
        except Exception as e:
            self.db.rollback()
            raise e

    # 2. Tạo Lệnh Sản Xuất (Thường) - GIỮ NGUYÊN
    def create_order(self, data: ProductionOrderCreateRequest):
        try:
            query = text("""
                INSERT INTO production_orders (code, warehouse_id, owner_central_id, product_variant_id, quantity_planned, status, start_date, due_date)
                VALUES (:code, :wid, :owner_central_id, :pid, :qty, 'draft', :start, :due)
            """)
            self.db.execute(query, {
                "code": data.code,
                "wid": data.warehouse_id,
                "owner_central_id": data.owner_central_id,
                "pid": data.product_variant_id,
                "qty": data.quantity_planned,
                "start": data.start_date,
                "due": data.due_date
            })
            self.db.commit()
            return {"status": "success", "message": f"Đã tạo lệnh sản xuất nháp: {data.code}"}
        except Exception as e:
            self.db.rollback()
            raise e

    def create_quick_order(self, data: QuickProductionRequest, user_id: int = None):
        try:
            # 0) Chặn trùng mã lệnh ngay từ đầu (tránh tạo product orphan)
            existed = self.db.execute(
                text("SELECT id FROM production_orders WHERE code = :code LIMIT 1"),
                {"code": data.order_code},
            ).fetchone()
            if existed:
                raise Exception("Mã lệnh đã tồn tại, vui lòng chọn mã khác")

            # A. Tính tổng số lượng sản phẩm
            if hasattr(data, 'size_breakdown') and data.size_breakdown:
                total_planned = sum(item.quantity for item in data.size_breakdown)
            else:
                total_planned = data.quantity_planned
            
            if total_planned <= 0: raise Exception("Tổng số lượng sản phẩm phải lớn hơn 0")

            # B. Tạo Sản phẩm & Variant
            query_prod = text("INSERT INTO products (category_id, name, type, base_unit) VALUES (3, :name, 'finished_good', 'Cái')")
            self.db.execute(query_prod, {"name": data.new_product_name})
            pid = self.db.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]

            query_var = text("""
                INSERT INTO product_variants (product_id, sku, variant_name, cost_price)
                VALUES (:pid, :sku, :name, 0)
            """)
            self.db.execute(query_var, {"pid": pid, "sku": data.new_product_sku, "name": data.new_product_name})
            product_variant_id = self.db.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]

            # C. Tạo BOM (Lưu định mức/cái)
            query_bom = text("INSERT INTO bom (product_variant_id, name) VALUES (:pid, :name)")
            self.db.execute(query_bom, {"pid": product_variant_id, "name": f"Công thức {data.new_product_name}"})
            bom_id = self.db.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]

            query_bom_detail = text("INSERT INTO bom_materials (bom_id, material_variant_id, quantity_needed, note) VALUES (:bid, :mid, :qty, :note)")
            
            total_material_cost = 0.0

            for item in data.materials:
                # Tính định mức cho 1 sản phẩm = Tổng lượng / Tổng sản phẩm
                # Ví dụ: Cần 1000m vải cho 500 áo -> Định mức = 2m/áo
                per_unit_usage = float(item.quantity_needed) / float(total_planned)

                self.db.execute(query_bom_detail, {
                    "bid": bom_id, 
                    "mid": item.material_variant_id, 
                    "qty": per_unit_usage, # Lưu định mức
                    "note": item.note if item.note else ""
                })

                # Tính giá vốn (Lấy giá * Tổng lượng)
                mat_price_row = self.db.execute(text("SELECT cost_price FROM product_variants WHERE id = :id"), {"id": item.material_variant_id}).fetchone()
                price = float(mat_price_row[0]) if (mat_price_row and mat_price_row[0] is not None) else 0.0
                total_material_cost += (float(item.quantity_needed) * price)

            # D. Tính giá vốn đơn vị
            total_fees = float(data.labor_fee + data.shipping_fee + data.other_fee + data.marketing_fee + data.packaging_fee + data.print_fee)
            final_unit_cost = (total_material_cost + total_fees) / float(total_planned)

            self.db.execute(text("UPDATE product_variants SET cost_price = :cost WHERE id = :id"), {
                "cost": final_unit_cost, "id": product_variant_id
            })


            default_steps = [
                {"name": "Bước 1: Chuẩn bị NVL & Rập", "done": False, "deadline": str(data.start_date)},
                {"name": "Bước 2: Cắt bán thành phẩm", "done": False, "deadline": str(data.start_date)}, # +3 ngày tùy logic
                {"name": "Bước 3: May gia công", "done": False, "deadline": str(data.due_date)},
                {"name": "Bước 4: KCS & Đóng gói", "done": False, "deadline": str(data.due_date)},
            ]
            progress_json = json.dumps(default_steps)


            # E. Tạo Lệnh SX
            query_order = text("""
                INSERT INTO production_orders (
                    code, warehouse_id, owner_central_id, product_variant_id, quantity_planned, status, start_date, due_date, 
                    shipping_fee, other_fee, labor_fee, marketing_fee, packaging_fee, print_fee, created_by, progress_data, note
                )
                VALUES (:code, :wid, :owner_central_id, :pid, :qty, 'draft', :start, :due, :ship, :other, :labor, :mkt, :pack, :print, :uid, :progress, :note)
            """)
            self.db.execute(query_order, {
                "code": data.order_code, "wid": data.warehouse_id, "pid": product_variant_id,
                "owner_central_id": data.owner_central_id,
                "qty": total_planned, "start": data.start_date, "due": data.due_date,
                "ship": data.shipping_fee, "other": data.other_fee, "labor": data.labor_fee,
                "mkt": data.marketing_fee, "pack": data.packaging_fee, "print": data.print_fee,
                "uid": user_id if user_id else None,
                "progress": progress_json,
                "note": data.note if data.note else ""
            })
            order_id = self.db.execute(text("SELECT LAST_INSERT_ID()")).fetchone()[0]

            # F. Lưu Size & Ảnh
            if hasattr(data, 'size_breakdown') and data.size_breakdown:
                query_size = text("INSERT INTO production_order_items (production_order_id, size_label, quantity_planned, quantity_finished, note) VALUES (:oid, :size, :qty, 0, :note)")
                for item in data.size_breakdown:
                    self.db.execute(query_size, {"oid": order_id, "size": item.size, "qty": item.quantity, "note": item.note if item.note else ""})

            if hasattr(data, 'image_urls') and data.image_urls:
                query_img = text("INSERT INTO production_order_images (production_order_id, image_url) VALUES (:oid, :url)")
                for url in data.image_urls:
                    self.db.execute(query_img, {"oid": order_id, "url": url})

            if data.auto_start:
                self.start_production(order_id, commit=False) 
            self.db.commit()

            return {"status": "success", "message": "Đã tạo Mẫu & Lệnh SX thành công!"}

        except IntegrityError as e:
            self.db.rollback()
            if "Duplicate entry" in str(e):
                # Sau migration 20260427_drop_unique_product_variants_sku.sql, SKU có thể trùng.
                # Duplicate còn lại chủ yếu là mã lệnh (production_orders.code).
                raise Exception("Lỗi: Mã lệnh đã tồn tại!")
            raise Exception(str(e))
        except Exception as e:
            self.db.rollback()
            raise e

    # 3. CẬP NHẬT TIẾN ĐỘ (PROGRESS) - YÊU CẦU MỚI
    def update_progress(self, order_id: int, data: UpdateProgressRequest):
        # Chuyển object thành string JSON để lưu
        json_str = json.dumps([item.dict() for item in data.steps])
        self.db.execute(text("UPDATE production_orders SET progress_data = :p WHERE id = :id"), {"p": json_str, "id": order_id})
        self.db.commit()
        return {"status": "success", "message": "Đã cập nhật tiến độ!"}


    # 4. Bắt đầu SX (Có tham số commit để hỗ trợ transaction lồng)
    def start_production(self, order_id: int, commit: bool = True):
        try:
            # Lấy thông tin lệnh
            order = self.db.execute(text("""
                SELECT warehouse_id, product_variant_id, quantity_planned
                FROM production_orders
                WHERE id = :id
            """), {"id": order_id}).fetchone()
            if not order: raise Exception("Không tìm thấy lệnh SX")
            wid, product_variant_id, quantity_planned = order
            bom_items = self.db.execute(text("""
                SELECT bm.material_variant_id, bm.quantity_needed 
                FROM bom_materials bm
                JOIN bom b ON bm.bom_id = b.id
                WHERE b.product_variant_id = :pid
            """), {"pid": product_variant_id}).fetchall()

            if not bom_items: raise Exception("Sản phẩm này chưa có công thức (BOM)!")

            # Duyệt qua từng nguyên liệu để TRỪ KHO
            for item in bom_items:
                mat_id = item[0]
                qty_needed_per_unit = float(item[1]) 
                order_qty = float(quantity_planned)
                raw_total = qty_needed_per_unit * order_qty
               
                total_qty_needed = round(raw_total, 4) 

                stock = self.db.execute(text("""
                    SELECT quantity_on_hand
                    FROM inventory_stocks 
                    WHERE warehouse_id = :wid AND product_variant_id = :mid
                """), {"wid": wid, "mid": mat_id}).fetchone()

                current_stock = float(stock[0]) if stock else 0.0

                if current_stock < total_qty_needed:
                    raise Exception(f"Kho thiếu nguyên liệu ID {mat_id}. Cần {total_qty_needed}, chỉ còn {current_stock}")

                # --- TRỪ TRỰC TIẾP ---
                self.db.execute(text("""
                    UPDATE inventory_stocks 
                    SET quantity_on_hand = quantity_on_hand - :qty
                    WHERE warehouse_id = :wid AND product_variant_id = :mid
                """), {"qty": total_qty_needed, "wid": wid, "mid": mat_id})

                # Ghi log xuất kho
                self.db.execute(text("""
                    INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id)
                    VALUES (:wid, :mid, 'production_out', :qty, :ref)
                """), {"wid": wid, "mid": mat_id, "qty": -total_qty_needed, "ref": order_id})

                # Lưu reservation
                self.db.execute(text("""
                    INSERT INTO production_material_reservations (production_order_id, material_variant_id, quantity_reserved)
                    VALUES (:oid, :mid, :qty)
                """), {"oid": order_id, "mid": mat_id, "qty": total_qty_needed})

            # Cập nhật trạng thái lệnh
            self.db.execute(text("UPDATE production_orders SET status = 'in_progress' WHERE id = :id"), {"id": order_id})
            
            if commit:
                self.db.commit()
                
            return {"status": "success", "message": "Đã xuất kho nguyên liệu & Bắt đầu sản xuất!"}

        except Exception as e:
            if commit: self.db.rollback()
            # Nếu gọi lồng, để hàm cha rollback
            raise e

    # 5. NHẬP KHO THÀNH PHẨM TỪNG ĐỢT (Receive Goods)
    def receive_goods(self, order_id: int, data: ReceiveGoodsRequest):
        try:
            # ... (Phần lấy thông tin lệnh giữ nguyên) ...
            order = self.db.execute(text("""
                SELECT po.warehouse_id, po.product_variant_id, po.code, pv.sku
                FROM production_orders po
                JOIN product_variants pv ON pv.id = po.product_variant_id
                WHERE po.id = :id
            """), {"id": order_id}).fetchone()
            if not order: raise Exception("Không tìm thấy lệnh")
            
            wid, pid, order_code, product_sku = order
            total_received_now = 0
            qty_by_inventory_code: Dict[str, float] = {}

            for item in data.items:
                if item.quantity <= 0: continue

                # A. Update số lượng đã xong (Giữ nguyên)
                if item.id:
                    self.db.execute(text("UPDATE production_order_items SET quantity_finished = quantity_finished + :qty WHERE id = :item_id"), {"qty": item.quantity, "item_id": item.id})
                    
                    # --- MỚI: GHI LOG LỊCH SỬ ---
                    self.db.execute(text("""
                        INSERT INTO production_receive_logs (production_order_id, production_order_item_id, quantity)
                        VALUES (:oid, :item_id, :qty)
                    """), {"oid": order_id, "item_id": item.id, "qty": item.quantity})
                    # ---------------------------
                else:
                    raise Exception("Thiếu ID dòng size")

                self.db.execute(text("""
                    INSERT INTO inventory_stocks (warehouse_id, product_variant_id, quantity_on_hand)
                    VALUES (:wid, :pid, :qty)
                    ON DUPLICATE KEY UPDATE quantity_on_hand = quantity_on_hand + :qty
                """), {"wid": wid, "pid": pid, "qty": item.quantity})

                total_received_now += item.quantity
                inv_code = (getattr(item, "inventory_code", None) or "").strip()
                if inv_code:
                    qty_by_inventory_code[inv_code] = float(qty_by_inventory_code.get(inv_code, 0.0)) + float(item.quantity)

            self.db.execute(text("UPDATE production_orders SET quantity_finished = quantity_finished + :qty WHERE id = :id"), {"qty": total_received_now, "id": order_id})
            
            self.db.execute(text("""
                INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id, note)
                VALUES (:wid, :pid, 'production_in', :qty, :ref, :note)
            """), {"wid": wid, "pid": pid, "qty": total_received_now, "ref": order_id, "note": f"Trả hàng đợt: {total_received_now}"})

            # --- NEW: Ghi tăng Kế toán theo mã kiểm tồn do người nhập điền (inventory_code) ---
            # - Nếu để trống: không lỗi, chỉ không ghi movement.
            # - Nếu có nhiều dòng size cùng 1 mã: sẽ cộng dồn theo mã.
            if qty_by_inventory_code and total_received_now > 0:
                # period_month theo kỳ kiểm tồn đang active (nếu chưa có table periods thì fallback current month)
                try:
                    period_row = self.db.execute(text("SELECT period_month FROM inventory_check_periods WHERE is_active=1 ORDER BY id DESC LIMIT 1")).fetchone()
                    period_month = str(period_row[0]) if period_row and period_row[0] else datetime.now().strftime("%Y-%m")
                except Exception:
                    period_month = datetime.now().strftime("%Y-%m")
                insert_move = text("""
                    INSERT INTO accounting_stock_movements
                        (period_month, product_code, movement_date, movement_type, direction, quantity, reason, document_ref, source_file)
                    VALUES
                        (:period, :code, CURDATE(), 'Nhập gia công', 'inc', :qty, :reason, :doc, 'production_receive')
                """)
                for inv_code, qty in qty_by_inventory_code.items():
                    if qty <= 0:
                        continue
                    self.db.execute(insert_move, {
                        "period": period_month,
                        "code": str(inv_code).strip(),
                        "qty": float(qty),
                        "reason": f"Nhập kho thành phẩm từ lệnh {order_code}",
                        "doc": f"PROD_RECEIVE:{order_id}",
                    })

            self.db.commit()
            return {"status": "success", "message": f"Đã nhập {total_received_now} SP và lưu lịch sử."}

        except Exception as e:
            self.db.rollback()
            raise e

    def get_receive_history(self, order_id: int):
        query = text("""
            SELECT 
                l.id,
                l.received_at,
                i.size_label,
                i.note as size_note,
                l.quantity,
                i.quantity_planned, 
                i.quantity_finished 
            FROM production_receive_logs l
            JOIN production_order_items i ON l.production_order_item_id = i.id
            WHERE l.production_order_id = :oid
            ORDER BY l.received_at DESC
        """)
        results = self.db.execute(query, {"oid": order_id}).fetchall()
        
        return [
            {
                "id": r[0],
                "date": r[1].strftime("%Y-%m-%d %H:%M"),
                "size": r[2],
                "note": r[3],
                "quantity": r[4],
                "remaining": r[5] - r[6]
            } for r in results
        ]

    # 6. HÀM CHỐT ĐƠN (Force Finish)
    def force_finish_order(self, order_id: int):
        # Không cho hoàn thành nếu trùng cả (mã lệnh + tên sản phẩm) với đơn khác
        row = self.db.execute(
            text("""
                SELECT po.code, pv.variant_name
                FROM production_orders po
                JOIN product_variants pv ON pv.id = po.product_variant_id
                WHERE po.id = :id
            """),
            {"id": order_id},
        ).fetchone()
        if not row:
            raise Exception("Không tìm thấy lệnh sản xuất")
        code, product_name = row[0], row[1]
        dup = self.db.execute(
            text("""
                SELECT COUNT(1)
                FROM production_orders po
                JOIN product_variants pv ON pv.id = po.product_variant_id
                WHERE po.id <> :id
                  AND po.code = :code
                  AND pv.variant_name = :name
            """),
            {"id": order_id, "code": code, "name": product_name},
        ).scalar() or 0
        if int(dup) > 0:
            raise Exception("Không thể hoàn thành: trùng cả mã lệnh và tên sản phẩm")

        self.db.execute(text("UPDATE production_orders SET status = 'completed' WHERE id = :id"), {"id": order_id})
        self.db.commit()
        return {"status": "success", "message": "Đã chốt hoàn thành đơn hàng."}

    # 7. API LẤY CHI TIẾT SIZE
    def get_order_details(self, order_id: int):
        results = self.db.execute(text("""
            SELECT id, size_label, quantity_planned, quantity_finished 
            FROM production_order_items 
            WHERE production_order_id = :oid
        """), {"oid": order_id}).fetchall()
        return [{"id": r[0], "size": r[1], "planned": r[2], "finished": r[3]} for r in results]

# 6. Lấy danh sách Lệnh SX (CHUẨN PHÂN TRANG)
    def get_all_orders(
        self,
        page=1,
        limit=10,
        search=None,
        warehouse_name=None,
        status=None,
        allowed_warehouse_ids: Optional[List[int]] = None,
        allowed_central_ids: Optional[List[int]] = None,
    ):
        offset = (page - 1) * limit
        
        conditions = []
        params = {"limit": limit, "offset": offset}

        if allowed_warehouse_ids is not None:
            if len(allowed_warehouse_ids) == 0:
                return {"data": [], "total": 0, "page": page, "limit": limit}

            ids_str = ",".join(map(str, allowed_warehouse_ids))
            conditions.append(f"po.warehouse_id IN ({ids_str})")

        if allowed_central_ids is not None:
            if len(allowed_central_ids) == 0:
                conditions.append("po.owner_central_id IS NULL")
            else:
                cids_str = ",".join(map(str, allowed_central_ids))
                conditions.append(f"(po.owner_central_id IS NULL OR po.owner_central_id IN ({cids_str}))")

        if search:
            conditions.append("(po.code LIKE :search OR pv.variant_name LIKE :search)")
            params["search"] = f"%{search}%"
        
        if warehouse_name:
            conditions.append("w.name = :wname")
            params["wname"] = warehouse_name
        if status:
            conditions.append("po.status = :status")
            params["status"] = status
        conditions.append("po.status != 'cancelled'")

        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

        # Đếm tổng
        count_sql = text(f"""
            SELECT COUNT(*)
            FROM production_orders po
            JOIN warehouses w ON po.warehouse_id = w.id
            JOIN product_variants pv ON po.product_variant_id = pv.id
            {where_clause}
        """)
        total_records = self.db.execute(count_sql, params).scalar()

        # Lấy dữ liệu
        data_sql = text(f"""
            SELECT po.id, po.code, w.name as warehouse_name, pv.variant_name as product_name,
                   po.quantity_planned, po.quantity_finished, po.status, po.start_date, po.due_date, po.progress_data, po.warehouse_id,
                   COALESCE(wc.brand_id, w.brand_id) as owner_brand_id
            FROM production_orders po
            JOIN warehouses w ON po.warehouse_id = w.id
            LEFT JOIN warehouses wc ON wc.id = po.owner_central_id
            JOIN product_variants pv ON po.product_variant_id = pv.id 
            {where_clause}
            ORDER BY
                CASE WHEN po.status = 'completed' THEN 1 ELSE 0 END ASC,
                w.name ASC,
                po.id DESC
            LIMIT :limit OFFSET :offset
        """)
        results = self.db.execute(data_sql, params).fetchall()

        result_list = []
        for r in results:
            # Xử lý Progress: Chuyển từ JSON String -> Python List/Dict
            progress_raw = r[9]
            progress_parsed = []
            if progress_raw:
                try:
                    if isinstance(progress_raw, str):
                        progress_parsed = json.loads(progress_raw)
                    else:
                        progress_parsed = progress_raw # Trường hợp DB driver tự parse
                except:
                    progress_parsed = [] # Nếu lỗi JSON thì trả về rỗng

            result_list.append({
                "id": r[0], 
                "code": r[1], 
                "warehouse_name": r[2], 
                "product_name": r[3],
                "quantity_planned": r[4], 
                "quantity_finished": r[5], 
                "status": r[6],
                "start_date": r[7], 
                "due_date": r[8], 
                "warehouse_id": r[10],
                "owner_brand_id": int(r[11]) if r[11] is not None else None,
                "progress": progress_parsed 
            })

        return {
            "data": result_list,
            "total": total_records,
            "page": page,
            "limit": limit
        }

    def export_orders_excel(
        self,
        start_date_from: Optional[date] = None,
        allowed_warehouse_ids: Optional[List[int]] = None,
        allowed_central_ids: Optional[List[int]] = None,
    ) -> bytes:
        """
        Export production orders to Excel (the Salework template columns).
        Filter: start_date >= start_date_from (only "start" time as requested).
        Sort: by workshop name.
        """
        conditions = ["po.status != 'cancelled'"]
        params: Dict[str, object] = {}

        if allowed_warehouse_ids is not None:
            if len(allowed_warehouse_ids) == 0:
                # Return empty workbook with headers
                allowed_warehouse_ids = []
            if len(allowed_warehouse_ids) == 0:
                conditions.append("1=0")
            else:
                ids_str = ",".join(map(str, allowed_warehouse_ids))
                conditions.append(f"po.warehouse_id IN ({ids_str})")

        if allowed_central_ids is not None:
            if len(allowed_central_ids) == 0:
                conditions.append("po.owner_central_id IS NULL")
            else:
                cids_str = ",".join(map(str, allowed_central_ids))
                conditions.append(f"(po.owner_central_id IS NULL OR po.owner_central_id IN ({cids_str}))")

        if start_date_from is not None:
            conditions.append("po.start_date >= :start_date_from")
            params["start_date_from"] = start_date_from

        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

        rows = self.db.execute(
            text(f"""
                SELECT
                    po.id,
                    po.code,
                    w.name AS workshop_name,
                    pv.sku AS product_sku,
                    pv.variant_name AS product_name,
                    po.start_date,
                    po.due_date,
                    po.status,
                    po.quantity_planned,
                    po.labor_fee,
                    po.print_fee,
                    po.shipping_fee,
                    po.marketing_fee,
                    po.packaging_fee,
                    po.other_fee,
                    b.name AS brand_name,
                    po.product_variant_id
                FROM production_orders po
                JOIN warehouses w ON w.id = po.warehouse_id
                LEFT JOIN warehouses wc ON wc.id = po.owner_central_id
                JOIN brands b ON b.id = COALESCE(wc.brand_id, w.brand_id)
                JOIN product_variants pv ON pv.id = po.product_variant_id
                {where_clause}
                ORDER BY w.name ASC, po.id DESC
            """),
            params,
        ).fetchall()

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # Header (match template text & merged structure)
        ws["A1"] = "STT"
        ws["B1"] = "Nhãn hàng"
        ws["C1"] = "Xưởng"
        ws["D1"] = "Mã lệnh"
        ws["E1"] = "Mã SKU"
        ws["F1"] = "Tên sản phẩm"
        ws["G1"] = "Thời gian"
        ws["I1"] = "Sản phẩm"
        ws["K1"] = "Nguyên phụ liệu"
        ws["N1"] = "Chi phí"
        ws["T1"] = "Chi phí chung"

        ws["G2"] = "Bắt đầu"
        ws["H2"] = "kết thúc"
        ws["I2"] = "Size"
        ws["J2"] = "Số lượng"
        ws["K2"] = "Tên"
        ws["L2"] = "Số lượng"
        ws["M2"] = "Giá trị"
        ws["N2"] = "Gia công"
        ws["O2"] = "In/ Thêu"
        ws["P2"] = "Vận chuyển"
        ws["Q2"] = "Marketing"
        ws["R2"] = "Đóng gói"
        ws["S2"] = "Phụ phí"

        # Merges (from template snapshot)
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:B2")
        ws.merge_cells("C1:C2")
        ws.merge_cells("D1:D2")
        ws.merge_cells("E1:E2")
        ws.merge_cells("F1:F2")
        ws.merge_cells("G1:H1")
        ws.merge_cells("I1:J1")
        ws.merge_cells("K1:M1")
        ws.merge_cells("N1:S1")
        ws.merge_cells("T1:T2")

        def _fmt_date(d: Optional[date]) -> str:
            if not d:
                return ""
            try:
                return d.strftime("%Y-%m-%d")
            except Exception:
                return str(d)

        out_row = 3
        stt = 1
        for (
            order_id,
            code,
            workshop_name,
            product_sku,
            product_name,
            start_date,
            due_date,
            status,
            quantity_planned,
            labor_fee,
            print_fee,
            shipping_fee,
            marketing_fee,
            packaging_fee,
            other_fee,
            brand_name,
            product_variant_id,
        ) in rows:
            # Sizes (tách theo dòng)
            size_rows = self.db.execute(
                text("""
                    SELECT size_label, quantity_planned
                    FROM production_order_items
                    WHERE production_order_id = :oid
                    ORDER BY id ASC
                """),
                {"oid": int(order_id)},
            ).fetchall()
            size_pairs: List[Dict[str, str]] = []
            for s_label, s_qty in size_rows:
                if s_label is None:
                    continue
                size_pairs.append(
                    {
                        "size": str(s_label),
                        "qty": str(int(s_qty or 0)),
                    }
                )

            # Materials (estimate or actual) (tách theo dòng)
            material_items: List[Dict[str, object]] = []
            material_total_value = 0.0

            if status in ("in_progress", "completed"):
                mats = self.db.execute(
                    text("""
                        SELECT m.variant_name, pmr.quantity_reserved, m.cost_price
                        FROM production_material_reservations pmr
                        JOIN product_variants m ON m.id = pmr.material_variant_id
                        WHERE pmr.production_order_id = :oid
                        ORDER BY pmr.id ASC
                    """),
                    {"oid": int(order_id)},
                ).fetchall()
                for m_name, m_qty, m_cost in mats:
                    q = float(m_qty or 0)
                    c = float(m_cost or 0)
                    v = q * c
                    material_total_value += v
                    material_items.append(
                        {
                            "name": str(m_name or ""),
                            "qty": q,
                            "value": v,
                        }
                    )
            else:
                mats = self.db.execute(
                    text("""
                        SELECT m.variant_name, bm.quantity_needed, m.cost_price
                        FROM bom b
                        JOIN bom_materials bm ON bm.bom_id = b.id
                        JOIN product_variants m ON m.id = bm.material_variant_id
                        WHERE b.product_variant_id = :pid
                        ORDER BY bm.id ASC
                    """),
                    {"pid": int(product_variant_id)},
                ).fetchall()
                total_qty = float(quantity_planned or 0)
                for m_name, usage, m_cost in mats:
                    q = float(usage or 0) * total_qty
                    c = float(m_cost or 0)
                    v = q * c
                    material_total_value += v
                    material_items.append(
                        {
                            "name": str(m_name or ""),
                            "qty": q,
                            "value": v,
                        }
                    )

            fees = [
                float(labor_fee or 0),
                float(print_fee or 0),
                float(shipping_fee or 0),
                float(marketing_fee or 0),
                float(packaging_fee or 0),
                float(other_fee or 0),
            ]
            fee_total = sum(fees)
            grand_total = material_total_value + fee_total

            lines = max(len(size_pairs), len(material_items), 1)
            for i in range(lines):
                sp = size_pairs[i] if i < len(size_pairs) else None
                mi = material_items[i] if i < len(material_items) else None

                ws[f"A{out_row}"] = stt if i == 0 else None
                ws[f"B{out_row}"] = brand_name if i == 0 else None
                ws[f"C{out_row}"] = workshop_name if i == 0 else None
                ws[f"D{out_row}"] = code if i == 0 else None
                ws[f"E{out_row}"] = product_sku if i == 0 else None
                ws[f"F{out_row}"] = product_name if i == 0 else None
                ws[f"G{out_row}"] = _fmt_date(start_date) if i == 0 else None
                ws[f"H{out_row}"] = _fmt_date(due_date) if i == 0 else None

                ws[f"I{out_row}"] = sp["size"] if sp else ""
                ws[f"J{out_row}"] = sp["qty"] if sp else ""
                ws[f"K{out_row}"] = mi["name"] if mi else ""
                # số lượng NVL có thể là lẻ -> vẫn để số, nhưng format không có phần thập phân theo yêu cầu hiện tại
                if mi:
                    ws[f"L{out_row}"] = float(mi.get("qty") or 0)
                    ws[f"M{out_row}"] = int(round(float(mi.get("value") or 0)))
                else:
                    ws[f"L{out_row}"] = ""
                    ws[f"M{out_row}"] = ""

                # Chi phí & tổng chỉ ghi ở dòng đầu của đơn (tránh bị nhân bản)
                if i == 0:
                    ws[f"N{out_row}"] = int(round(float(labor_fee or 0)))
                    ws[f"O{out_row}"] = int(round(float(print_fee or 0)))
                    ws[f"P{out_row}"] = int(round(float(shipping_fee or 0)))
                    ws[f"Q{out_row}"] = int(round(float(marketing_fee or 0)))
                    ws[f"R{out_row}"] = int(round(float(packaging_fee or 0)))
                    ws[f"S{out_row}"] = int(round(float(other_fee or 0)))
                    ws[f"T{out_row}"] = int(round(float(grand_total or 0)))

                # Format số để tránh scientific notation (e+06) và bỏ phần thập phân
                for col in ("M", "N", "O", "P", "Q", "R", "S", "T"):
                    cell = ws[f"{col}{out_row}"]
                    if isinstance(cell.value, (int, float)) and cell.value != "":
                        cell.number_format = "#,##0"

                out_row += 1

            stt += 1

        buf = BytesIO()
        wb.save(buf)
        return buf.getvalue()
    

    def get_orders_management(
        self,
        page: int = 1,
        limit: int = 50,
        search: Optional[str] = None,
        warehouse_id: Optional[int] = None,
        start_date_from: Optional[date] = None,
        start_date_to: Optional[date] = None,
        due_date_from: Optional[date] = None,
        due_date_to: Optional[date] = None,
        include_completed: bool = False,
        allowed_warehouse_ids: Optional[List[int]] = None,
        allowed_central_ids: Optional[List[int]] = None,
        allowed_material_cost_brand_ids: Optional[List[int]] = None,
    ):
        """
        "Quản lý đơn": trả danh sách đơn sản xuất + size + nguyên phụ liệu + chi phí để render bảng ngang.
        - Mặc định: ẩn completed/cancelled.
        - Sort: theo xưởng (workshop) để kế toán/kiểm kê dễ quan sát.
        """
        offset = (page - 1) * limit

        conditions = ["po.status != 'cancelled'"]
        params: Dict[str, object] = {"limit": limit, "offset": offset}

        if not include_completed:
            conditions.append("po.status != 'completed'")

        if allowed_warehouse_ids is not None:
            if len(allowed_warehouse_ids) == 0:
                conditions.append("1=0")
            else:
                ids_str = ",".join(map(str, allowed_warehouse_ids))
                conditions.append(f"po.warehouse_id IN ({ids_str})")

        if allowed_central_ids is not None:
            if len(allowed_central_ids) == 0:
                conditions.append("po.owner_central_id IS NULL")
            else:
                cids_str = ",".join(map(str, allowed_central_ids))
                conditions.append(f"(po.owner_central_id IS NULL OR po.owner_central_id IN ({cids_str}))")

        if warehouse_id is not None:
            conditions.append("po.warehouse_id = :warehouse_id")
            params["warehouse_id"] = warehouse_id

        if search:
            conditions.append("(po.code LIKE :search OR pv.sku LIKE :search OR pv.variant_name LIKE :search)")
            params["search"] = f"%{search}%"

        if start_date_from is not None:
            conditions.append("po.start_date >= :start_date_from")
            params["start_date_from"] = start_date_from
        if start_date_to is not None:
            conditions.append("po.start_date <= :start_date_to")
            params["start_date_to"] = start_date_to
        if due_date_from is not None:
            conditions.append("po.due_date >= :due_date_from")
            params["due_date_from"] = due_date_from
        if due_date_to is not None:
            conditions.append("po.due_date <= :due_date_to")
            params["due_date_to"] = due_date_to

        where_clause = "WHERE " + " AND ".join(conditions) if conditions else ""

        total_records = (
            self.db.execute(
                text(
                    f"""
                    SELECT COUNT(*)
                    FROM production_orders po
                    JOIN warehouses w ON w.id = po.warehouse_id
                    LEFT JOIN warehouses wc ON wc.id = po.owner_central_id
                    LEFT JOIN brands b ON b.id = COALESCE(wc.brand_id, w.brand_id)
                    JOIN product_variants pv ON pv.id = po.product_variant_id
                    {where_clause}
                    """
                ),
                params,
            ).scalar()
            or 0
        )

        orders = self.db.execute(
            text(
                f"""
                SELECT
                    po.id,
                    po.code,
                    po.status,
                    po.progress_data,
                    po.warehouse_id,
                    w.name AS workshop_name,
                    b.id AS brand_id,
                    b.name AS brand_name,
                    pv.id AS product_variant_id,
                    pv.sku AS product_sku,
                    pv.variant_name AS product_name,
                    pv.cost_price AS product_cost_price,
                    po.quantity_planned,
                    po.quantity_finished,
                    po.start_date,
                    po.due_date,
                    po.labor_fee,
                    po.print_fee,
                    po.shipping_fee,
                    po.marketing_fee,
                    po.packaging_fee,
                    po.other_fee
                FROM production_orders po
                JOIN warehouses w ON w.id = po.warehouse_id
                LEFT JOIN warehouses wc ON wc.id = po.owner_central_id
                LEFT JOIN brands b ON b.id = COALESCE(wc.brand_id, w.brand_id)
                JOIN product_variants pv ON pv.id = po.product_variant_id
                {where_clause}
                ORDER BY w.name ASC, po.id DESC
                LIMIT :limit OFFSET :offset
                """
            ),
            params,
        ).fetchall()

        order_ids = [int(r[0]) for r in orders]

        # Sizes (all orders in one query)
        sizes_by_order: Dict[int, List[dict]] = {oid: [] for oid in order_ids}
        if order_ids:
            ids_str = ",".join(map(str, order_ids))
            size_rows = self.db.execute(
                text(
                    f"""
                    SELECT production_order_id, size_label, quantity_planned, quantity_finished
                    FROM production_order_items
                    WHERE production_order_id IN ({ids_str})
                    ORDER BY production_order_id ASC, id ASC
                    """
                )
            ).fetchall()
            for oid, size_label, qty_planned, qty_finished in size_rows:
                sizes_by_order[int(oid)].append(
                    {
                        "size": size_label,
                        "quantity": float(qty_planned or 0),
                        "finished": float(qty_finished or 0),
                    }
                )

        # Materials:
        mats_by_order: Dict[int, List[dict]] = {oid: [] for oid in order_ids}

        in_progress_or_completed = set()
        for r in orders:
            oid, status = int(r[0]), r[2]
            if status in ["in_progress", "completed"]:
                in_progress_or_completed.add(oid)

        if order_ids:
            # Reservations path
            if in_progress_or_completed:
                ids_str = ",".join(map(str, sorted(in_progress_or_completed)))
                res_rows = self.db.execute(
                    text(
                        f"""
                        SELECT
                            pmr.production_order_id,
                            pv.sku,
                            pv.variant_name,
                            pmr.quantity_reserved,
                            pv.cost_price,
                            pmr.note
                        FROM production_material_reservations pmr
                        JOIN product_variants pv ON pv.id = pmr.material_variant_id
                        WHERE pmr.production_order_id IN ({ids_str})
                        ORDER BY pmr.production_order_id ASC, pmr.id ASC
                        """
                    )
                ).fetchall()
                for oid, sku, name, qty, unit_cost, note in res_rows:
                    qty_f = float(Decimal(str(qty or 0)).quantize(Decimal("0.0001")))
                    unit = float(unit_cost or 0)
                    mats_by_order[int(oid)].append(
                        {
                            "sku": sku,
                            "name": name,
                            "quantity": qty_f,
                            "unit_cost": unit,
                            "total_cost": qty_f * unit,
                            "note": note,
                        }
                    )

            # BOM path for draft/waiting_material
            draft_ids = [oid for oid in order_ids if oid not in in_progress_or_completed]
            if draft_ids:
                ids_str = ",".join(map(str, draft_ids))
                bom_rows = self.db.execute(
                    text(
                        f"""
                        SELECT
                            po.id AS production_order_id,
                            m.sku,
                            m.variant_name,
                            bm.quantity_needed,
                            po.quantity_planned,
                            m.cost_price,
                            bm.note
                        FROM production_orders po
                        JOIN bom b ON b.product_variant_id = po.product_variant_id
                        JOIN bom_materials bm ON bm.bom_id = b.id
                        JOIN product_variants m ON m.id = bm.material_variant_id
                        WHERE po.id IN ({ids_str})
                        ORDER BY po.id ASC, bm.id ASC
                        """
                    )
                ).fetchall()
                for oid, sku, name, per_unit, qty_planned, unit_cost, note in bom_rows:
                    total_needed = float(
                        (Decimal(str(per_unit or 0)) * Decimal(str(qty_planned or 0))).quantize(Decimal("0.0001"))
                    )
                    unit = float(unit_cost or 0)
                    mats_by_order[int(oid)].append(
                        {
                            "sku": sku,
                            "name": name,
                            "quantity": total_needed,
                            "unit_cost": unit,
                            "total_cost": total_needed * unit,
                            "note": note,
                        }
                    )

        # Apply cost visibility by brand
        allowed_brand_set = None
        if allowed_material_cost_brand_ids is not None:
            allowed_brand_set = set(int(x) for x in allowed_material_cost_brand_ids)

        data_out = []
        for r in orders:
            (
                order_id,
                code,
                status,
                progress_data,
                wid,
                workshop_name,
                brand_id,
                brand_name,
                product_variant_id,
                product_sku,
                product_name,
                product_cost_price,
                qty_planned,
                qty_finished,
                start_date,
                due_date,
                labor_fee,
                print_fee,
                shipping_fee,
                marketing_fee,
                packaging_fee,
                other_fee,
            ) = r

            owner_brand_id = int(brand_id) if brand_id is not None else None
            progress_parsed = []
            if progress_data:
                try:
                    progress_parsed = json.loads(progress_data) if isinstance(progress_data, str) else progress_data
                except Exception:
                    progress_parsed = []
            can_view_cost = (
                allowed_brand_set is None
                or (owner_brand_id is not None and owner_brand_id in allowed_brand_set)
            )

            mats = mats_by_order.get(int(order_id), [])
            if not can_view_cost:
                mats = [{**m, "unit_cost": None, "total_cost": None} for m in mats]

            data_out.append(
                {
                    "id": int(order_id),
                    "code": code,
                    "status": status,
                    "progress": progress_parsed if isinstance(progress_parsed, list) else [],
                    "warehouse_id": int(wid) if wid is not None else None,
                    "workshop_name": workshop_name,
                    "brand_id": owner_brand_id,
                    "brand_name": brand_name,
                    "product_variant_id": int(product_variant_id) if product_variant_id is not None else None,
                    "product_sku": product_sku,
                    "product_name": product_name,
                    "product_cost_price": float(product_cost_price) if (product_cost_price is not None and can_view_cost) else None,
                    "quantity_planned": float(qty_planned or 0),
                    "quantity_finished": float(qty_finished or 0),
                    "start_date": start_date,
                    "due_date": due_date,
                    "sizes": sizes_by_order.get(int(order_id), []),
                    "materials": mats,
                    "fees": {
                        "labor_fee": float(labor_fee or 0) if can_view_cost else None,
                        "print_fee": float(print_fee or 0) if can_view_cost else None,
                        "shipping_fee": float(shipping_fee or 0) if can_view_cost else None,
                        "marketing_fee": float(marketing_fee or 0) if can_view_cost else None,
                        "packaging_fee": float(packaging_fee or 0) if can_view_cost else None,
                        "other_fee": float(other_fee or 0) if can_view_cost else None,
                    },
                    "can_view_cost": can_view_cost,
                }
            )

        return {"data": data_out, "total": int(total_records), "page": page, "limit": limit}

    def get_all_boms(self):
        query = text("""
            SELECT b.id, b.name, p.variant_name as product_name
            FROM bom b
            JOIN product_variants p ON b.product_variant_id = p.id
            ORDER BY b.id DESC
        """)
        results = self.db.execute(query).fetchall()
        return [{"id": r[0], "name": r[1], "product_name": r[2]} for r in results]
    
    # ... (Các hàm cũ giữ nguyên) ...

    # 8. LẤY DỮ LIỆU ĐỂ IN (ĐÃ CẬP NHẬT LOGIC LẤY NVL THỰC TẾ)
    def get_order_print_data(self, order_id: int):
        # A. Header (Giữ nguyên)
        query_header = text("""
            SELECT po.code, w.name as warehouse_name, w.address as warehouse_address,
                   pv.variant_name as product_name, pv.sku as product_sku,
                   po.quantity_planned, po.start_date, po.due_date,
                   po.product_variant_id,
                   po.shipping_fee, po.other_fee,
                   po.labor_fee, po.marketing_fee, po.packaging_fee, po.print_fee, 
                   po.status,
                   po.note,
                   COALESCE(wc.brand_id, w.brand_id) as owner_brand_id
            FROM production_orders po
            JOIN warehouses w ON po.warehouse_id = w.id
            JOIN product_variants pv ON po.product_variant_id = pv.id
            LEFT JOIN warehouses wc ON wc.id = po.owner_central_id
            WHERE po.id = :oid
        """)
        header = self.db.execute(query_header, {"oid": order_id}).fetchone()
        if not header: raise Exception("Không tìm thấy lệnh")
        
        status = header[15]
        total_qty = header[5]

        # B. Size (Giữ nguyên)
        query_sizes = text("SELECT size_label, quantity_planned, note FROM production_order_items WHERE production_order_id = :oid ORDER BY id ASC")
        sizes = self.db.execute(query_sizes, {"oid": order_id}).fetchall()
        list_sizes = [{"size": s[0], "qty": s[1], "note": s[2]} for s in sizes]

        # C. Materials (LOGIC MỚI: Ưu tiên lấy từ Reservation nếu có)
        list_materials = []
        total_material_cost = 0

        # Nếu đơn đã Start hoặc có Reservation -> Lấy từ bảng Reservation (Chính xác nhất sau khi sửa)
        if status in ['in_progress', 'completed']:
            query_res = text("""
                SELECT m.sku, m.variant_name, pmr.quantity_reserved, m.cost_price, pmr.note
                FROM production_material_reservations pmr
                JOIN product_variants m ON pmr.material_variant_id = m.id
                WHERE pmr.production_order_id = :oid
            """)
            materials = self.db.execute(query_res, {"oid": order_id}).fetchall()
            
            for m in materials:
                # Sửa dòng này: Thêm round(..., 4)
                total_needed = round(float(m[2]), 4)
                
                unit_cost = float(m[3] or 0)
                total_cost_mat = total_needed * unit_cost
                total_material_cost += total_cost_mat

                list_materials.append({
                    "sku": m[0], "name": m[1], 
                    "total_needed": total_needed, # Giá trị này giờ đã tròn đẹp (ví dụ 26.2)
                    "unit_cost": unit_cost,
                    "total_cost": total_cost_mat,
                    "note": m[4]
                })
        
        # Nếu chưa Start -> Lấy từ BOM (Dự kiến)
        else:
            query_bom = text("""
                SELECT m.sku, m.variant_name, bm.quantity_needed, m.cost_price, bm.note
                FROM bom b
                JOIN bom_materials bm ON b.id = bm.bom_id
                JOIN product_variants m ON bm.material_variant_id = m.id
                WHERE b.product_variant_id = :pid
            """)
            materials = self.db.execute(query_bom, {"pid": header[8]}).fetchall()
            
            for m in materials:
                usage = float(m[2])
                total_needed = usage * total_qty # Nhân định mức với số lượng
                unit_cost = float(m[3] or 0)
                total_cost_mat = total_needed * unit_cost
                total_material_cost += total_cost_mat

                list_materials.append({
                    "sku": m[0], "name": m[1], 
                    "total_needed": total_needed,
                    "unit_cost": unit_cost,
                    "total_cost": total_cost_mat,
                    "note": m[4]
                })

        # D. Ảnh (Giữ nguyên)
        query_imgs = text("SELECT image_url FROM production_order_images WHERE production_order_id = :oid")
        list_imgs = [r[0] for r in self.db.execute(query_imgs, {"oid": order_id}).fetchall()]

        return {
            "code": header[0],
            "warehouse": header[1],
            "address": header[2],
            "product": header[3],
            "sku": header[4],
            "total_qty": total_qty,
            "start_date": header[6],
            "due_date": header[7],
            
            "shipping_fee": header[9],
            "other_fee": header[10],
            "labor_fee": header[11],
            "marketing_fee": header[12],
            "packaging_fee": header[13],
            "print_fee": header[14],
            "note": header[16] if header[16] else "",
            "owner_brand_id": int(header[17]) if header[17] is not None else None,
            "total_material_cost": total_material_cost,
            "sizes": list_sizes,
            "materials": list_materials,
            "images": list_imgs
        }
    
    def update_production_order(self, order_id: int, data: ProductionUpdateRequest):
        try:
            print(f"--- DEBUG UPDATE ORDER {order_id} ---")
            order = self.db.execute(text("SELECT warehouse_id, status, product_variant_id, quantity_planned FROM production_orders WHERE id = :id"), {"id": order_id}).fetchone()
            if not order: raise Exception("Không tìm thấy đơn hàng")
            wid = order[0]
            status = order[1]
            pid = order[2]
            old_qty_planned = Decimal(str(order[3])) # Chuyển sang Decimal ngay

            # 2. Cập nhật Header
            query_header = text("""
                UPDATE production_orders
                SET shipping_fee = :ship, other_fee = :other, labor_fee = :labor, print_fee = :print,
                    marketing_fee = :mkt, packaging_fee = :pack,
                    start_date = :start, due_date = :due, note = :note
                WHERE id = :id
            """)
            self.db.execute(query_header, {
                "ship": data.shipping_fee, "other": data.other_fee, "labor": data.labor_fee, "print": data.print_fee,
                "mkt": data.marketing_fee, "pack": data.packaging_fee,
                "start": data.start_date, "due": data.due_date, "id": order_id, "note": data.note if data.note else ""
            })

            total_qty_planned = Decimal("0")
            
            if data.sizes is not None:
                existing_items = self.db.execute(text("SELECT id FROM production_order_items WHERE production_order_id = :oid"), {"oid": order_id}).fetchall()
                existing_ids = [r[0] for r in existing_items]
                sent_ids = [s.id for s in data.sizes if s.id]
                ids_to_delete = set(existing_ids) - set(sent_ids)
                
                for del_id in ids_to_delete:
                    finished = self.db.execute(text("SELECT quantity_finished FROM production_order_items WHERE id = :id"), {"id": del_id}).scalar()
                    if finished and finished > 0:
                        raise Exception(f"Không thể xóa Size (ID {del_id}) vì đã có sản phẩm hoàn thành!")
                    
                    self.db.execute(text("DELETE FROM production_order_items WHERE id = :id"), {"id": del_id})
                # -------------------------

                # --- LOGIC THÊM / SỬA ---
                for s in data.sizes:
                    qty_decimal = Decimal(str(s.quantity))
                    total_qty_planned += qty_decimal
                    
                    if s.id:
                        self.db.execute(text("UPDATE production_order_items SET size_label=:size, quantity_planned=:qty, note=:note WHERE id=:id"),
                                        {"size": s.size, "qty": qty_decimal, "note": s.note, "id": s.id})
                    else:
                        # Thêm size mới
                        self.db.execute(text("INSERT INTO production_order_items (production_order_id, size_label, quantity_planned, quantity_finished, note) VALUES (:oid, :size, :qty, 0, :note)"),
                                        {"oid": order_id, "size": s.size, "qty": qty_decimal, "note": s.note})
                self.db.execute(text("UPDATE production_orders SET quantity_planned = :q WHERE id = :id"), {"q": total_qty_planned, "id": order_id})

            current_qty_planned = total_qty_planned if (data.sizes is not None and total_qty_planned > 0) else old_qty_planned

            # 3. XỬ LÝ NGUYÊN VẬT LIỆU
            if data.materials is not None:
                print(f"Materials count: {len(data.materials)}")
                for item in data.materials:
                    req_qty = Decimal(str(item.quantity))
                    if status in ['in_progress', 'completed']:
                        print("-> Branch: IN_PROGRESS")
                        
                        if item.id:
                            old_res = self.db.execute(text("SELECT quantity_reserved, material_variant_id FROM production_material_reservations WHERE id=:id"), {"id": item.id}).fetchone()
                            if old_res:
                                old_qty = Decimal(str(old_res[0]))
                                mat_id = old_res[1]
                                
                                diff = req_qty - old_qty

                                if diff > 0:
                                    stock = self.db.execute(text("SELECT quantity_on_hand FROM inventory_stocks WHERE warehouse_id=:w AND product_variant_id=:m"), {"w": wid, "m": mat_id}).scalar() or 0
                                    stock_dec = Decimal(str(stock))
                                    
                                    if stock_dec < diff: 
                                        raise Exception(f"Kho không đủ hàng! Cần thêm {diff}, còn {stock_dec}")
                                    
                                    self.db.execute(text("UPDATE inventory_stocks SET quantity_on_hand = quantity_on_hand - :q WHERE warehouse_id=:w AND product_variant_id=:m"), {"q": diff, "w": wid, "m": mat_id})
                                    self.db.execute(text("INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id, note) VALUES (:w, :m, 'production_out', :q, :ref, 'Sửa lệnh: Cấp thêm')"), {"w": wid, "m": mat_id, "q": -diff, "ref": order_id})
                                
                                elif diff < 0:
                                    return_qty = abs(diff)
                                    self.db.execute(text("UPDATE inventory_stocks SET quantity_on_hand = quantity_on_hand + :q WHERE warehouse_id=:w AND product_variant_id=:m"), {"q": return_qty, "w": wid, "m": mat_id})
                                    self.db.execute(text("INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id, note) VALUES (:w, :m, 'production_in', :q, :ref, 'Sửa lệnh: Hoàn trả')"), {"w": wid, "m": mat_id, "q": return_qty, "ref": order_id})

                                # Nếu người dùng xóa NVL (set về 0) thì xóa luôn reservation để sạch kế hoạch
                                if req_qty <= 0:
                                    self.db.execute(
                                        text("DELETE FROM production_material_reservations WHERE id = :id"),
                                        {"id": item.id},
                                    )
                                else:
                                    self.db.execute(
                                        text("UPDATE production_material_reservations SET quantity_reserved = :q, note = :n WHERE id = :id"),
                                        {"q": req_qty, "n": item.note, "id": item.id},
                                    )
                        else: 
                            # Thêm mới
                            if not item.material_variant_id: continue
                            mat_id = item.material_variant_id
                            
                            stock = self.db.execute(text("SELECT quantity_on_hand FROM inventory_stocks WHERE warehouse_id=:w AND product_variant_id=:m"), {"w": wid, "m": mat_id}).scalar() or 0
                            stock_dec = Decimal(str(stock))

                            if stock_dec < req_qty: 
                                raise Exception(f"Kho không đủ hàng mới! Cần {req_qty}, còn {stock_dec}")
                            
                            self.db.execute(text("UPDATE inventory_stocks SET quantity_on_hand = quantity_on_hand - :q WHERE warehouse_id=:w AND product_variant_id=:m"), {"q": req_qty, "w": wid, "m": mat_id})
                            self.db.execute(text("INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id, note) VALUES (:w, :m, 'production_out', :q, :ref, 'Thêm NVL mới')"), {"w": wid, "m": mat_id, "q": -req_qty, "ref": order_id})
                            
                            self.db.execute(text("INSERT INTO production_material_reservations (production_order_id, material_variant_id, quantity_reserved, note) VALUES (:oid, :mid, :qty, :note)"), 
                                            {"oid": order_id, "mid": mat_id, "qty": req_qty, "note": item.note})

                    # === TRƯỜNG HỢP B: ĐƠN NHÁP (DRAFT) -> Sửa bảng BOM ===
                    else:
                        print("-> Branch: DRAFT (BOM)")
                        bom_id = self.db.execute(text("SELECT id FROM bom WHERE product_variant_id=:pid"), {"pid": pid}).scalar()
                        
                        if bom_id:
                            print(f"   Calculating: RequestQty={req_qty} / PlannedQty={current_qty_planned}")
                            
                            # Tính định mức bằng Decimal
                            per_unit_qty = req_qty / current_qty_planned if current_qty_planned > 0 else Decimal("0")
                            
                            if item.id:
                                print(f"   Updating BOM Material ID={item.id}")
                                if req_qty <= 0:
                                    # Xóa dòng BOM nếu user xóa NVL khỏi kế hoạch
                                    self.db.execute(text("DELETE FROM bom_materials WHERE id = :id"), {"id": item.id})
                                else:
                                    self.db.execute(text("UPDATE bom_materials SET quantity_needed = :q, note = :n WHERE id = :id"), 
                                                    {"q": per_unit_qty, "n": item.note, "id": item.id})
                            else:
                                if item.material_variant_id:
                                    if req_qty > 0:
                                        self.db.execute(text("INSERT INTO bom_materials (bom_id, material_variant_id, quantity_needed, note) VALUES (:bid, :mid, :q, :n)"),
                                                        {"bid": bom_id, "mid": item.material_variant_id, "q": per_unit_qty, "n": item.note})

            # 4. Cập nhật SKU
            if (hasattr(data, 'new_sku') and data.new_sku) or (hasattr(data, 'new_product_name') and data.new_product_name):
                # Lấy PID
                pid = self.db.execute(text("SELECT product_variant_id FROM production_orders WHERE id=:id"), {"id": order_id}).scalar()
                
                # Cập nhật bảng Variants
                update_fields = {}
                if data.new_sku: update_fields['sku'] = data.new_sku
                if data.new_product_name: update_fields['name'] = data.new_product_name
                
                # Xây dựng câu SQL động
                set_clause = ", ".join([f"variant_name = :{k}" if k=='name' else f"{k} = :{k}" for k in update_fields.keys()])
                if set_clause:
                    update_fields['pid'] = pid
                    self.db.execute(text(f"UPDATE product_variants SET {set_clause} WHERE id = :pid"), update_fields)
                    
                    # Nếu muốn cập nhật luôn tên bảng cha (products) cho đồng bộ
                    if data.new_product_name:
                        parent_id = self.db.execute(text("SELECT product_id FROM product_variants WHERE id=:pid"), {"pid": pid}).scalar()
                        self.db.execute(text("UPDATE products SET name = :name WHERE id = :pid"), {"name": data.new_product_name, "pid": parent_id})


            # 5. Cập nhật Ảnh
            if data.image_urls is not None:
                self.db.execute(text("DELETE FROM production_order_images WHERE production_order_id = :oid"), {"oid": order_id})
                if len(data.image_urls) > 0:
                    query_img = text("INSERT INTO production_order_images (production_order_id, image_url) VALUES (:oid, :url)")
                    for url in data.image_urls:
                        self.db.execute(query_img, {"oid": order_id, "url": url})

            self.db.commit()
            return {"status": "success", "message": "Cập nhật thành công!"}
        except Exception as e:
            self.db.rollback()
            raise e
        


    # 8. XÓA ĐƠN HÀNG & HOÀN KHO & XÓA SKU TẠO NHANH
    def delete_production_order(self, order_id: int):
        try:
            # A. Lấy thông tin lệnh
            order = self.db.execute(text("SELECT status, warehouse_id, product_variant_id, quantity_finished, code FROM production_orders WHERE id = :id"), {"id": order_id}).fetchone()
            if not order: raise Exception("Đơn hàng không tồn tại")

            status = order[0] 
            warehouse_id = order[1]
            product_variant_id = order[2] 
            qty_finished = float(order[3]) if order[3] else 0.0
            order_code = order[4]

            if status == 'cancelled':
                raise Exception("Đơn hàng này đã bị hủy từ trước.")

            # B. Hoàn trả NVL & Thành phẩm (Chỉ áp dụng nếu đơn đã chạy)
            if status in ['in_progress', 'completed']:
                # 1. Hoàn trả NVL đang giữ chỗ
                reservations = self.db.execute(text("""
                    SELECT material_variant_id, quantity_reserved 
                    FROM production_material_reservations 
                    WHERE production_order_id = :oid
                """), {"oid": order_id}).fetchall()

                for res in reservations:
                    mat_id = res[0]
                    qty_return = float(res[1]) if res[1] else 0.0
                    
                    if qty_return > 0:
                        # Cộng lại tồn kho
                        self.db.execute(text("""
                            UPDATE inventory_stocks 
                            SET quantity_on_hand = quantity_on_hand + :qty
                            WHERE warehouse_id = :wid AND product_variant_id = :mid
                        """), {"qty": qty_return, "wid": warehouse_id, "mid": mat_id})

                        # Ghi log hoàn trả
                        self.db.execute(text("""
                            INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id, note)
                            VALUES (:wid, :mid, 'production_in', :qty, :ref, 'Hoàn trả NVL do Hủy đơn')
                        """), {"wid": warehouse_id, "mid": mat_id, "qty": qty_return, "ref": order_id})

                # 2. Hoàn trả (Trừ đi) Thành phẩm đã nhập kho (nếu có)
                if qty_finished > 0:
                     self.db.execute(text("""
                        UPDATE inventory_stocks SET quantity_on_hand = quantity_on_hand - :qty
                        WHERE warehouse_id = :wid AND product_variant_id = :pid
                     """), {"qty": qty_finished, "wid": warehouse_id, "pid": product_variant_id})
                     
                     self.db.execute(text("""
                        INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id, note)
                        VALUES (:wid, :pid, 'production_out', :qty, :ref, 'Hủy nhập TP do Hủy đơn')
                    """), {"wid": warehouse_id, "pid": product_variant_id, "qty": -qty_finished, "ref": order_id})

            # C. ĐỔI TRẠNG THÁI VÀ GIẢI PHÓNG MÃ (CHỐNG TRÙNG LẶP)
            # Tạo chuỗi duy nhất từ thời gian hiện tại
            timestamp_suffix = f"_C_{int(time.time())}"
            
            # Cắt ngắn mã cũ để không vượt quá giới hạn VARCHAR(50)
            short_order_code = str(order_code)[:35]
            new_code = f"{short_order_code}{timestamp_suffix}"
            
            # Đổi mã lệnh và chuyển trạng thái
            self.db.execute(text("UPDATE production_orders SET status = 'cancelled', code = :new_code WHERE id = :id"), 
                            {"new_code": new_code, "id": order_id})

            # Đổi mã SKU sản phẩm
            old_sku = self.db.execute(text("SELECT sku FROM product_variants WHERE id = :pid"), {"pid": product_variant_id}).scalar()
            if old_sku:
                short_sku = str(old_sku)[:35]
                new_sku = f"{short_sku}{timestamp_suffix}"
                self.db.execute(text("UPDATE product_variants SET sku = :new_sku WHERE id = :pid"), 
                                {"new_sku": new_sku, "pid": product_variant_id})

            # Đưa số lượng giữ chỗ (reservations) về 0 để số liệu báo cáo không bị ảo
            self.db.execute(text("UPDATE production_material_reservations SET quantity_reserved = 0 WHERE production_order_id = :id"), {"id": order_id})

            self.db.commit()
            return {"status": "success", "message": "Đã Hủy đơn hàng và hoàn kho thành công. Bạn có thể dùng lại mã này."}

        except Exception as e:
            self.db.rollback()
            raise e

    def revert_receive_log(self, log_id: int):
        try:
            # 1. Lấy thông tin log nhập hàng cũ
            log = self.db.execute(text("""
                SELECT l.production_order_id, l.production_order_item_id, l.quantity, 
                       po.warehouse_id, po.product_variant_id
                FROM production_receive_logs l
                JOIN production_orders po ON l.production_order_id = po.id
                WHERE l.id = :lid
            """), {"lid": log_id}).fetchone()

            if not log:
                raise Exception("Không tìm thấy lịch sử nhập này")

            order_id, item_id, qty, wid, pid = log

            # 2. Trừ kho thành phẩm (Revert Stock)
            current_stock = self.db.execute(text("SELECT quantity_on_hand FROM inventory_stocks WHERE warehouse_id=:w AND product_variant_id=:p"), {"w": wid, "p": pid}).scalar() or 0
            if current_stock < qty:
                raise Exception(f"Không thể hoàn tác! Kho chỉ còn {current_stock} sản phẩm (Cần trừ {qty}).")

            self.db.execute(text("""
                UPDATE inventory_stocks 
                SET quantity_on_hand = quantity_on_hand - :qty
                WHERE warehouse_id = :wid AND product_variant_id = :pid
            """), {"qty": qty, "wid": wid, "pid": pid})

            # 3. Trừ số lượng đã hoàn thành trong Order Header & Item
            self.db.execute(text("UPDATE production_orders SET quantity_finished = quantity_finished - :qty WHERE id = :id"), {"qty": qty, "id": order_id})
            self.db.execute(text("UPDATE production_order_items SET quantity_finished = quantity_finished - :qty WHERE id = :id"), {"qty": qty, "id": item_id})

            # 4. Ghi log kho (Transaction Out để cân bằng)
            self.db.execute(text("""
                INSERT INTO inventory_transactions (warehouse_id, product_variant_id, transaction_type, quantity, reference_id, note)
                VALUES (:wid, :pid, 'production_out', :qty, :ref, :note)
            """), {"wid": wid, "pid": pid, "qty": -qty, "ref": order_id, "note": f"Hoàn tác nhập kho #{log_id}"})

            # 5. Xóa dòng log
            self.db.execute(text("DELETE FROM production_receive_logs WHERE id = :id"), {"id": log_id})

            self.db.commit()
            return {"status": "success", "message": "Đã hoàn tác nhập kho thành công."}
        except Exception as e:
            self.db.rollback()
            raise e

            
    # CẬP NHẬT THÊM ID VÀO HÀM LẤY LỊCH SỬ ĐỂ FRONTEND GỌI XÓA ĐƯỢC
    def get_receive_history(self, order_id: int):
        query = text("""
            SELECT 
                l.id,  -- Thêm lấy ID log
                l.received_at,
                i.size_label,
                i.note as size_note,
                l.quantity,
                i.quantity_planned, 
                i.quantity_finished
            FROM production_receive_logs l
            JOIN production_order_items i ON l.production_order_item_id = i.id
            WHERE l.production_order_id = :oid
            ORDER BY l.received_at DESC
        """)
        results = self.db.execute(query, {"oid": order_id}).fetchall()
        
        return [
            {
                "id": r[0], # Trả về ID
                "date": r[1].strftime("%Y-%m-%d %H:%M"),
                "size": r[2],
                "note": r[3],
                "quantity": r[4],
                "remaining": r[5] - r[6]
            } for r in results
        ]

    def get_order_reservations(self, order_id: int):
        # 1. Lấy trạng thái đơn hàng
        order = self.db.execute(text("SELECT status, product_variant_id, quantity_planned FROM production_orders WHERE id=:id"), {"id": order_id}).fetchone()
        if not order: return []
        
        status = order[0]
        pid = order[1]
        qty_planned = order[2]

        # TRƯỜNG HỢP 1: ĐÃ START (Lấy từ bảng Reservation - Dữ liệu thực tế đã trừ kho)
        if status in ['in_progress', 'completed']:
            query = text("""
                SELECT pmr.id, pmr.material_variant_id, pv.sku, pv.variant_name, pmr.quantity_reserved, pmr.note, pv.cost_price
                FROM production_material_reservations pmr
                JOIN product_variants pv ON pmr.material_variant_id = pv.id
                WHERE pmr.production_order_id = :oid
            """)
            results = self.db.execute(query, {"oid": order_id}).fetchall()
            return [
                {
                    "id": r[0], 
                    "material_variant_id": r[1], 
                    "sku": r[2], 
                    "name": r[3], 
                    "quantity": float(Decimal(str(r[4] or 0)).quantize(Decimal("0.0001"))), 
                    "note": r[5], 
                    "unit_price": r[6] or 0
                } for r in results
            ]

        # TRƯỜNG HỢP 2: CHƯA START (Lấy từ BOM - Dữ liệu lý thuyết)
        else:
            query = text("""
                SELECT bm.id, bm.material_variant_id, pv.sku, pv.variant_name, bm.quantity_needed, bm.note, pv.cost_price
                FROM bom_materials bm
                JOIN bom b ON bm.bom_id = b.id
                JOIN product_variants pv ON bm.material_variant_id = pv.id
                WHERE b.product_variant_id = :pid
            """)
            results = self.db.execute(query, {"pid": pid}).fetchall()
            
            return [
                {
                    "id": r[0],
                    "material_variant_id": r[1],
                    "sku": r[2], "name": r[3],
                    
                    # --- SỬA ĐOẠN NÀY: Dùng Decimal để nhân chính xác rồi mới làm tròn ---
                    "quantity": float(
                        (Decimal(str(r[4])) * Decimal(str(qty_planned))).quantize(Decimal("0.0001"))
                    ),
                    # -------------------------------------------------------------------
                    
                    "note": r[5], "unit_price": r[6] or 0
                } for r in results
            ]

          
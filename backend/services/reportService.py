# Báo cáo tồn kho
# Báo cáo nhập–xuất–tồn
# Báo cáo theo SKU
# Báo cáo sản xuất

import io
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from sqlalchemy.orm import Session
from sqlalchemy import text, bindparam
from urllib.parse import quote
from typing import Optional, List


class ReportService:
    def __init__(self, db: Session):
        self.db = db
        self._ensure_central_workshop_links_table()

    def _ensure_central_workshop_links_table(self):
        self.db.execute(text("""
            CREATE TABLE IF NOT EXISTS central_workshop_links (
                central_warehouse_id INT NOT NULL,
                workshop_warehouse_id INT NOT NULL,
                PRIMARY KEY (central_warehouse_id, workshop_warehouse_id),
                CONSTRAINT fk_cwl_central FOREIGN KEY (central_warehouse_id) REFERENCES warehouses(id) ON DELETE CASCADE,
                CONSTRAINT fk_cwl_workshop FOREIGN KEY (workshop_warehouse_id) REFERENCES warehouses(id) ON DELETE CASCADE
            )
        """))
        self.db.commit()

    def _get_dashboard_warehouse_ids(self, central_warehouse_id: int, brand_id: int):
        linked_rows = self.db.execute(text("""
            SELECT workshop_warehouse_id
            FROM central_workshop_links
            WHERE central_warehouse_id = :cid
        """), {"cid": central_warehouse_id}).fetchall()
        linked_workshops = {int(r[0]) for r in linked_rows}
        # luôn include dữ liệu theo brand cũ để tránh mất xưởng hiện tại
        old_rows = self.db.execute(
            text("SELECT id FROM warehouses WHERE brand_id = :bid"),
            {"bid": brand_id},
        ).fetchall()
        old_ids = {int(r[0]) for r in old_rows}
        return list({int(central_warehouse_id)} | old_ids | linked_workshops)

    def get_central_warehouse_dashboard(self, warehouse_id: int, visible_warehouse_ids: Optional[List[int]] = None):
        # 1. Lấy thông tin Kho Tổng & Brand của nó
        query_info = text("""
            SELECT w.id, w.name, w.brand_id, b.name as brand_name 
            FROM warehouses w
            JOIN brands b ON w.brand_id = b.id
            WHERE w.id = :wid AND w.is_central = 1
        """)
        info = self.db.execute(query_info, {"wid": warehouse_id}).fetchone()
        
        if not info:
            raise Exception("Không tìm thấy Kho Tổng hoặc đây không phải là Kho Tổng!")
        
        brand_id = int(info[2])
        dashboard_wids = self._get_dashboard_warehouse_ids(warehouse_id, brand_id)
        if visible_warehouse_ids is not None:
            visible_set = {int(x) for x in visible_warehouse_ids}
            dashboard_wids = [wid for wid in dashboard_wids if int(wid) in visible_set]
            # vẫn giữ kho tổng hiện tại nếu user có quyền vào endpoint này
            if int(warehouse_id) not in dashboard_wids:
                dashboard_wids = [int(warehouse_id)] + dashboard_wids

        # 2. Lấy danh sách Xưởng Con (Các kho cùng Brand nhưng không phải Central)
        query_children = text("""
            SELECT id, name, address
            FROM warehouses
            WHERE id IN :ids AND is_central = 0
            ORDER BY name ASC
        """).bindparams(bindparam("ids", expanding=True))
        children = self.db.execute(query_children, {"ids": dashboard_wids}).fetchall()
        list_children = [{"id": r[0], "name": r[1], "address": r[2]} for r in children]

        # 3. Tổng hợp Tồn kho Toàn hệ thống (Kho Tổng + Tất cả Xưởng con)
        # Group by theo Mã hàng để biết tổng lượng vải/cúc của cả Brand đang là bao nhiêu
        query_total_stock = text("""
            SELECT pv.sku, pv.variant_name, p.base_unit,
                   SUM(s.quantity_on_hand) as total_qty,
                   SUM(s.quantity_on_hand * pv.cost_price) as total_value,
                   pv.note -- <--- Thêm note
            FROM inventory_stocks s
            JOIN warehouses w ON s.warehouse_id = w.id
            JOIN product_variants pv ON s.product_variant_id = pv.id
            JOIN products p ON pv.product_id = p.id
            WHERE s.warehouse_id IN :ids
            GROUP BY pv.id, pv.sku, pv.variant_name, p.base_unit, pv.note -- Group by cả note
            HAVING total_qty > 0
        """).bindparams(bindparam("ids", expanding=True))
        stocks = self.db.execute(query_total_stock, {"ids": dashboard_wids}).fetchall()
        list_stocks = [
            {
                "sku": r[0], 
                "name": r[1], 
                "unit": r[2], 
                "total_quantity": r[3], 
                "total_value": r[4],
                "note": r[5]  # <--- Lấy note
            } for r in stocks
        ]

        # 4. Danh sách Đơn nhập hàng (Chỉ của Kho Tổng này)
        query_pos = text("""
            SELECT po.po_code, s.name as supplier_name, po.order_date, po.total_amount, po.status
            FROM purchase_orders po
            JOIN suppliers s ON po.supplier_id = s.id
            WHERE po.warehouse_id = :wid
            ORDER BY po.order_date DESC
            LIMIT 10 -- Lấy 10 đơn gần nhất
        """)
        pos = self.db.execute(query_pos, {"wid": warehouse_id}).fetchall()
        list_pos = [
            {
                "code": r[0], 
                "supplier": r[1], 
                "date": r[2], 
                "amount": r[3], 
                "status": r[4]
            } for r in pos
        ]

        # 5. Tình hình Sản xuất tại các Xưởng con (Các lệnh đang chạy)
        # Để biết xưởng nào đang may cái gì, bao nhiêu
        query_production = text("""
            SELECT po.code, w.name as workshop_name, pv.variant_name as product_name,
                   po.quantity_planned, po.quantity_finished, po.status, po.due_date
            FROM production_orders po
            JOIN warehouses w ON po.warehouse_id = w.id
            JOIN product_variants pv ON po.product_variant_id = pv.id
            WHERE po.warehouse_id IN :ids
            AND po.status IN ('in_progress', 'draft') -- Chỉ quan tâm cái đang làm
            ORDER BY po.due_date ASC
        """).bindparams(bindparam("ids", expanding=True))
        productions = self.db.execute(query_production, {"ids": dashboard_wids}).fetchall()
        list_production = [
            {
                "code": r[0],
                "workshop": r[1],
                "product": r[2],
                "planned": r[3],
                "finished": r[4],
                "status": r[5],
                "due_date": r[6]
            } for r in productions
        ]

        # Trả về cục dữ liệu tổng hợp
        return {
            "info": {"id": info[0], "name": info[1], "brand": info[3]},
            "workshops": list_children,
            "total_inventory": list_stocks,
            "recent_purchases": list_pos,
            "active_production": list_production
        }
    
    def get_workshop_detail(self, warehouse_id: int):
        # A. Thông tin cơ bản
        info = self.db.execute(text("SELECT id, name, address FROM warehouses WHERE id = :wid"), {"wid": warehouse_id}).fetchone()
        if not info: raise Exception("Không tìm thấy kho")

        # B. Tồn kho tại xưởng (Kèm giá trị)
        query_stock = text("""
            SELECT pv.sku, pv.variant_name, p.base_unit, 
                   s.quantity_on_hand, 
                   (s.quantity_on_hand * pv.cost_price) as total_value,
                   p.type, pv.note -- <--- Thêm note
            FROM inventory_stocks s
            JOIN product_variants pv ON s.product_variant_id = pv.id
            JOIN products p ON pv.product_id = p.id
            WHERE s.warehouse_id = :wid AND s.quantity_on_hand > 0
            ORDER BY total_value DESC
        """)
        stocks = self.db.execute(query_stock, {"wid": warehouse_id}).fetchall()
        
        list_stocks = [{
            "sku": r[0], "name": r[1], "unit": r[2], 
            "qty": r[3], "value": r[4], "type": r[5], "note": r[6]
        } for r in stocks]

        # C. Các đơn hàng đang may tại xưởng này
        query_orders = text("""
            SELECT code, pv.variant_name, quantity_planned, quantity_finished, status, due_date
            FROM production_orders po
            JOIN product_variants pv ON po.product_variant_id = pv.id
            WHERE po.warehouse_id = :wid
            ORDER BY po.id DESC
        """)
        orders = self.db.execute(query_orders, {"wid": warehouse_id}).fetchall()
        
        list_orders = [{
            "code": r[0], "product": r[1], 
            "planned": r[2], "finished": r[3], 
            "status": r[4], "due_date": r[5]
        } for r in orders]

        return {
            "info": {"id": info[0], "name": info[1], "address": info[2]},
            "inventory": list_stocks,
            "production": list_orders,
            "total_asset_value": sum(item['value'] for item in list_stocks)
        }
    
    def export_inventory_excel(self, central_warehouse_id: int, visible_warehouse_ids: Optional[List[int]] = None):
        # 1. Lấy thông tin Brand của Kho Tổng này
        info = self.db.execute(text("SELECT brand_id, name FROM warehouses WHERE id = :id AND is_central = 1"), {"id": central_warehouse_id}).fetchone()
        if not info:
            raise Exception("Không tìm thấy Kho Tổng!")
        brand_id = int(info[0])
        central_name = info[1]
        dashboard_wids = self._get_dashboard_warehouse_ids(central_warehouse_id, brand_id)
        if visible_warehouse_ids is not None:
            visible_set = {int(x) for x in visible_warehouse_ids}
            dashboard_wids = [wid for wid in dashboard_wids if int(wid) in visible_set]
            if int(central_warehouse_id) not in dashboard_wids:
                dashboard_wids = [int(central_warehouse_id)] + dashboard_wids

        # 2. Lấy danh sách toàn bộ tồn kho của Brand này (Cả Kho Tổng + Xưởng Con)
        # Bỏ qua những mã có tồn kho = 0 để file không bị rác
        query = text("""
            SELECT w.name as warehouse_name, 
                   pv.sku, 
                   pv.variant_name, 
                   p.base_unit, 
                   s.quantity_on_hand,
                   pv.cost_price,
                   (s.quantity_on_hand * pv.cost_price) as total_value,
                   pv.note
            FROM inventory_stocks s
            JOIN warehouses w ON s.warehouse_id = w.id
            JOIN product_variants pv ON s.product_variant_id = pv.id
            JOIN products p ON pv.product_id = p.id
            WHERE s.warehouse_id IN :ids AND s.quantity_on_hand > 0
            ORDER BY w.is_central DESC, w.name ASC, pv.variant_name ASC
        """).bindparams(bindparam("ids", expanding=True))
        stocks = self.db.execute(query, {"ids": dashboard_wids}).fetchall()

        # 3. Tạo file Excel bằng openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "BaoCaoTonKho"

        # --- Định dạng style ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        center_align = Alignment(horizontal="center", vertical="center")

        # --- Viết Tiêu đề báo cáo ---
        ws.merge_cells('A1:G1')
        ws['A1'] = f"BÁO CÁO TỒN KHO NGUYÊN PHỤ LIỆU - {central_name.upper()}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = center_align

        # --- Viết Header các cột ---
        headers = ["Kho/Xưởng", "Mã SKU", "Tên Nguyên Phụ Liệu", "ĐVT", "Tồn Kho", "Đơn Giá Vốn", "Thành Tiền", "Ghi chú"]
        ws.append([]) # Dòng trống
        ws.append(headers)
        
        # Format Header row (Dòng 3)
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=3, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align

        # --- Đổ dữ liệu vào các dòng ---
        for row_data in stocks:
            ws.append([
                row_data[0], # Kho/Xưởng
                row_data[1], # SKU
                row_data[2], # Tên
                row_data[3], # ĐVT
                float(row_data[4]), # Tồn
                float(row_data[5]), # Giá
                float(row_data[6]), # Tổng tiền
                row_data[7] or ""   # Note
            ])

        # --- Căn chỉnh độ rộng cột cho đẹp ---
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 15
        ws.column_dimensions['G'].width = 18
        ws.column_dimensions['H'].width = 20

        # 4. Lưu file Excel vào bộ nhớ đệm (RAM) để trả về thẳng HTTP Response
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Trả về file
        # Use RFC 5987 filename* for Unicode names; keep ASCII fallback for compatibility.
        encoded_name = quote(f"TonKho_{central_name.replace(' ', '')}.xlsx")
        headers = {
            "Content-Disposition": f"attachment; filename=\"TonKho.xlsx\"; filename*=UTF-8''{encoded_name}"
        }
        return StreamingResponse(output, headers=headers, media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')